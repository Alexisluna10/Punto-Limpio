from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.db.models import Sum, Count, Q
from decimal import Decimal
import json
from datetime import datetime, timedelta
from django.utils import timezone
from django.template.loader import get_template
from io import BytesIO
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.core.mail import EmailMessage
from django.conf import settings

# --- IMPORTACIONES DEL PROYECTO ---
from apps.core.decorators import solo_admin
from apps.usuarios.models import Usuario
from apps.servicios.models import Pedido, DetallePedido, MovimientoOperador
from apps.finanzas.models import CorteCaja, GastoRenta, GastoServicio, SalarioEmpleado
from django.views.decorators.http import require_POST

@solo_admin
def admin_finanzas(request):
    from django.db.models import F
    from calendar import monthrange

    hoy = timezone.now().date()

    # Determinar el periodo de filtro
    filtro = request.GET.get('filtro', 'hoy')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    tab_activa = request.GET.get('tab', 'resumen')

    if filtro == 'hoy':
        fecha_inicio = hoy
        fecha_fin = hoy
    elif filtro == 'semana':
        fecha_inicio = hoy - timedelta(days=7)
        fecha_fin = hoy
    elif filtro == 'mes':
        fecha_inicio = hoy.replace(day=1)
        fecha_fin = hoy
    elif filtro == 'personalizado' and fecha_desde and fecha_hasta:
        fecha_inicio = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    else:
        fecha_inicio = hoy
        fecha_fin = hoy

    # Calcular dias del periodo
    dias_periodo = (fecha_fin - fecha_inicio).days + 1

    # ========== DATOS FINANCIEROS - INGRESOS ==========
    pedidos_periodo = Pedido.objects.filter(
        fecha_recepcion__date__gte=fecha_inicio,
        fecha_recepcion__date__lte=fecha_fin,
        estado_pago='pagado'
    )

    ingresos_totales = pedidos_periodo.aggregate(
        total=Sum('total'))['total'] or Decimal('0')

    total_pedidos = pedidos_periodo.count()
    promedio_pedido = ingresos_totales / \
        total_pedidos if total_pedidos > 0 else Decimal('0')

    # ========== METODOS DE PAGO ==========
    pago_efectivo = pedidos_periodo.filter(metodo_pago='efectivo').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    pago_tarjeta = pedidos_periodo.filter(metodo_pago='tarjeta').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    pago_transferencia = pedidos_periodo.filter(metodo_pago='transferencia').aggregate(
        total=Sum('total'))['total'] or Decimal('0')

    total_pagos = pago_efectivo + pago_tarjeta + pago_transferencia
    pct_efectivo = round((pago_efectivo / total_pagos * 100),
                         1) if total_pagos > 0 else 0
    pct_tarjeta = round((pago_tarjeta / total_pagos * 100),
                        1) if total_pagos > 0 else 0
    pct_transferencia = round(
        (pago_transferencia / total_pagos * 100), 1) if total_pagos > 0 else 0

    # ========== GRAFICA DE PRENDAS ==========
    detalles_periodo = DetallePedido.objects.filter(
        pedido__fecha_recepcion__date__gte=fecha_inicio,
        pedido__fecha_recepcion__date__lte=fecha_fin,
        pedido__estado_pago='pagado'
    )

    prendas_stats = detalles_periodo.values(
        'prenda__nombre'
    ).annotate(
        cantidad_total=Sum('cantidad'),
        ganancia_total=Sum('subtotal')
    ).order_by('-cantidad_total')[:10]

    total_prendas = sum(p['cantidad_total']
                        for p in prendas_stats if p['cantidad_total']) if prendas_stats else 0
    prendas_data = []
    for prenda in prendas_stats:
        if prenda['prenda__nombre'] and prenda['cantidad_total']:
            pct = round(
                (prenda['cantidad_total'] / total_prendas * 100), 1) if total_prendas > 0 else 0
            prendas_data.append({
                'nombre': prenda['prenda__nombre'],
                'cantidad': prenda['cantidad_total'],
                'ganancia': float(prenda['ganancia_total'] or 0),
                'porcentaje': pct
            })

    # ========== GRAFICA DE SERVICIOS ==========
    servicios_stats = pedidos_periodo.values(
        'tipo_servicio'
    ).annotate(
        cantidad=Count('id'),
        ganancia_total=Sum('total')
    ).order_by('-cantidad')

    total_servicios = sum(s['cantidad']
                          for s in servicios_stats) if servicios_stats else 0
    servicios_data = []
    for servicio in servicios_stats:
        pct = round((servicio['cantidad'] / total_servicios *
                    100), 1) if total_servicios > 0 else 0
        servicios_data.append({
            'nombre': servicio['tipo_servicio'] or 'Sin especificar',
            'cantidad': servicio['cantidad'],
            'ganancia': float(servicio['ganancia_total'] or 0),
            'porcentaje': pct
        })

    # ========== DATOS DE GASTOS ==========
    mes_actual = hoy.month
    anio_actual = hoy.year

    # Nombres de los meses
    meses_nombres = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                     'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    meses = [{'num': i, 'nombre': meses_nombres[i-1]} for i in range(1, 13)]
    mes_nombre = meses_nombres[mes_actual - 1]

    # Obtener renta actual del mes
    renta_actual = GastoRenta.objects.filter(
        mes=mes_actual, anio=anio_actual).first()
    total_renta_mes = float(renta_actual.monto_mensual) if renta_actual else 0

    # Obtener servicios del mes
    tipos_servicio_data = [
        {'tipo': 'agua', 'nombre': 'Agua', 'icon': 'A'},
        {'tipo': 'luz', 'nombre': 'Luz', 'icon': 'L'},
        {'tipo': 'gas', 'nombre': 'Gas', 'icon': 'G'},
        {'tipo': 'internet', 'nombre': 'Internet', 'icon': 'I'},
    ]

    gastos_servicios_mes = GastoServicio.objects.filter(
        mes=mes_actual, anio=anio_actual)
    total_servicios_mes = Decimal('0')
    gastos_servicios_data = []

    for ts in tipos_servicio_data:
        gasto = gastos_servicios_mes.filter(tipo=ts['tipo']).first()
        ts['monto_actual'] = float(gasto.monto_mensual) if gasto else 0
        total_servicios_mes += Decimal(str(ts['monto_actual']))
        if ts['monto_actual'] > 0:
            gastos_servicios_data.append({
                'tipo': ts['nombre'],
                'monto': ts['monto_actual']
            })

    # Obtener empleados y sus salarios
    empleados = Usuario.objects.filter(rol__in=['admin', 'operador'])
    empleados_data = []
    total_sueldos_semanal = Decimal('0')
    sueldos_data = []

    for emp in empleados:
        salario = SalarioEmpleado.objects.filter(empleado=emp).first()
        salario_semanal = float(salario.salario_semanal) if salario else 0
        salario_diario = salario_semanal / 7 if salario_semanal > 0 else 0

        empleados_data.append({
            'id': emp.id,
            'username': emp.username,
            'first_name': emp.first_name,
            'last_name': emp.last_name,
            'rol': emp.rol,
            'salario_semanal': salario_semanal,
            'salario_diario': salario_diario,
        })
        total_sueldos_semanal += Decimal(str(salario_semanal))

        if salario_semanal > 0:
            sueldos_data.append({
                'nombre': emp.username,
                'rol': emp.rol,
                'sueldo_semanal': salario_semanal,
                'sueldo_diario': salario_diario
            })

    # Calcular sueldos mensuales (4 semanas)
    total_sueldos_mes = float(total_sueldos_semanal * 4)

    # Total gastos del mes
    total_gastos_mes = total_renta_mes + \
        float(total_servicios_mes) + total_sueldos_mes

    # ========== CALCULAR GASTOS PROPORCIONALES AL PERIODO ==========
    # Gastos diarios
    gasto_renta_diario = total_renta_mes / 30 if total_renta_mes > 0 else 0
    gasto_servicios_diario = float(
        total_servicios_mes) / 30 if total_servicios_mes > 0 else 0
    gasto_sueldos_diario = float(
        total_sueldos_semanal) / 7 if total_sueldos_semanal > 0 else 0

    # Gastos del periodo
    gastos_renta_periodo = gasto_renta_diario * dias_periodo
    gastos_servicios_periodo = gasto_servicios_diario * dias_periodo
    gastos_sueldos_periodo = gasto_sueldos_diario * dias_periodo
    total_gastos_periodo = gastos_renta_periodo + \
        gastos_servicios_periodo + gastos_sueldos_periodo

    # Utilidad neta
    utilidad_neta = float(ingresos_totales) - total_gastos_periodo

    # JSON para gastos
    gastos_json = {
        'renta': total_renta_mes,
        'servicios': gastos_servicios_data,
        'sueldos': total_sueldos_mes
    }

    context = {
        'filtro': filtro,
        'tab_activa': tab_activa,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'dias_periodo': dias_periodo,

        # Ingresos
        'ingresos_totales': float(ingresos_totales),
        'total_pedidos': total_pedidos,
        'promedio_pedido': float(promedio_pedido),
        'utilidad_neta': float(utilidad_neta),

        # Metodos de pago
        'pago_efectivo': float(pago_efectivo),
        'pago_tarjeta': float(pago_tarjeta),
        'pago_transferencia': float(pago_transferencia),
        'pct_efectivo': float(pct_efectivo),
        'pct_tarjeta': float(pct_tarjeta),
        'pct_transferencia': float(pct_transferencia),

        # JSON para graficas
        'prendas_json': json.dumps(prendas_data),
        'servicios_json': json.dumps(servicios_data),
        'metodos_pago_json': json.dumps([
            {'nombre': 'Efectivo', 'total': float(
                pago_efectivo), 'porcentaje': float(pct_efectivo)},
            {'nombre': 'Tarjeta', 'total': float(
                pago_tarjeta), 'porcentaje': float(pct_tarjeta)},
            {'nombre': 'Transferencia', 'total': float(
                pago_transferencia), 'porcentaje': float(pct_transferencia)},
        ]),
        'gastos_json': json.dumps(gastos_json),
        'sueldos_json': json.dumps(sueldos_data),

        # Datos para tab de gastos
        'meses': meses,
        'mes_actual': mes_actual,
        'mes_nombre': mes_nombre,
        'anio_actual': anio_actual,
        'renta_actual': renta_actual,
        'tipos_servicio': tipos_servicio_data,
        'empleados': empleados_data,

        # Totales mensuales
        'total_renta_mes': total_renta_mes,
        'total_servicios_mes': float(total_servicios_mes),
        'total_sueldos_mes': total_sueldos_mes,
        'total_gastos_mes': total_gastos_mes,

        # Gastos del periodo
        'gastos_renta_periodo': gastos_renta_periodo,
        'gastos_servicios_periodo': gastos_servicios_periodo,
        'gastos_sueldos_periodo': gastos_sueldos_periodo,
        'total_gastos_periodo': total_gastos_periodo,
    }
    return render(request, 'admin/finanzas/finanzas.html', context)

def admin_corte_caja(request):
    if not request.user.groups.filter(name='Administrador').exists():
        return redirect('core:tasks')

    # Obtener fecha de hoy
    hoy = timezone.now().date()

    # Verificar si ya existe un corte para hoy
    corte_existente = CorteCaja.objects.filter(
        fecha=hoy, responsable=request.user).first()

    # Si es POST, guardar el corte
    if request.method == 'POST':
        efectivo_contado = Decimal(request.POST.get('efectivo_contado', 0))
        tarjeta_terminal = Decimal(request.POST.get('tarjeta_terminal', 0))
        transferencia_banco = Decimal(
            request.POST.get('transferencia_banco', 0))
        justificacion = request.POST.get('justificacion', '')

        # Pedidos pagados del día de hoy
        pedidos_hoy = Pedido.objects.filter(
            fecha_recepcion__date=hoy,
            estado_pago='pagado'
        )

        # Calcular ventas por método de pago
        ventas_efectivo = pedidos_hoy.filter(metodo_pago='efectivo').aggregate(
            total=Sum('total'))['total'] or Decimal('0')
        ventas_tarjeta = pedidos_hoy.filter(metodo_pago='tarjeta').aggregate(
            total=Sum('total'))['total'] or Decimal('0')
        ventas_transferencia = pedidos_hoy.filter(metodo_pago='transferencia').aggregate(
            total=Sum('total'))['total'] or Decimal('0')

        total_ventas = ventas_efectivo + ventas_tarjeta + ventas_transferencia
        total_fisico = efectivo_contado + tarjeta_terminal + transferencia_banco
        diferencia = total_fisico - total_ventas

        # Crear o actualizar el corte
        if corte_existente:
            corte = corte_existente
            corte.efectivo_contado = efectivo_contado
            corte.tarjeta_terminal = tarjeta_terminal
            corte.transferencia_banco = transferencia_banco
            corte.total_fisico = total_fisico
            corte.diferencia = diferencia
            corte.justificacion = justificacion
            corte.fecha_hora_registro = timezone.now()
            messages.success(request, 'Corte de caja actualizado exitosamente')
        else:
            corte = CorteCaja(
                fecha=hoy,
                responsable=request.user,
                ventas_efectivo=ventas_efectivo,
                ventas_tarjeta=ventas_tarjeta,
                ventas_transferencia=ventas_transferencia,
                total_ventas=total_ventas,
                efectivo_contado=efectivo_contado,
                tarjeta_terminal=tarjeta_terminal,
                transferencia_banco=transferencia_banco,
                total_fisico=total_fisico,
                diferencia=diferencia,
                justificacion=justificacion
            )
            messages.success(request, 'Corte de caja guardado exitosamente')

        corte.save()

        # Registrar movimiento del operador
        MovimientoOperador.objects.create(
            operador=request.user,
            accion='actualizo',
            detalles=f'Corte de caja - Diferencia: ${diferencia}'
        )

        return redirect('finanzas:admin_corte_caja')

    # Pedidos pagados del día de hoy
    pedidos_hoy = Pedido.objects.filter(
        fecha_recepcion__date=hoy,
        estado_pago='pagado'
    )

    # Calcular ventas por método de pago
    ventas_efectivo = pedidos_hoy.filter(metodo_pago='efectivo').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    ventas_tarjeta = pedidos_hoy.filter(metodo_pago='tarjeta').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    ventas_transferencia = pedidos_hoy.filter(metodo_pago='transferencia').aggregate(
        total=Sum('total'))['total'] or Decimal('0')

    total_ventas = ventas_efectivo + ventas_tarjeta + ventas_transferencia

    # Si existe un corte, usar esos datos
    if corte_existente:
        efectivo_contado = corte_existente.efectivo_contado
        tarjeta_terminal = corte_existente.tarjeta_terminal
        transferencia_banco = corte_existente.transferencia_banco
        total_fisico = corte_existente.total_fisico
        diferencia = corte_existente.diferencia
        justificacion = corte_existente.justificacion or ''
    else:
        # Valores por defecto (vacios)
        efectivo_contado = Decimal('0')
        tarjeta_terminal = Decimal('0')
        transferencia_banco = Decimal('0')
        total_fisico = Decimal('0')
        diferencia = Decimal('0')
        justificacion = ''

    # ========== CALCULAR GASTOS DEL DIA ==========
    mes_actual = hoy.month
    anio_actual = hoy.year

    # Obtener renta del mes
    renta_actual = GastoRenta.objects.filter(
        mes=mes_actual, anio=anio_actual).first()
    gasto_renta_diario = float(
        renta_actual.monto_mensual) / 30 if renta_actual else 0

    # Obtener servicios del mes
    gastos_servicios_mes = GastoServicio.objects.filter(
        mes=mes_actual, anio=anio_actual)

    gasto_agua = gastos_servicios_mes.filter(tipo='agua').first()
    gasto_agua_diario = float(
        gasto_agua.monto_mensual) / 30 if gasto_agua else 0

    gasto_luz = gastos_servicios_mes.filter(tipo='luz').first()
    gasto_luz_diario = float(gasto_luz.monto_mensual) / 30 if gasto_luz else 0

    gasto_gas = gastos_servicios_mes.filter(tipo='gas').first()
    gasto_gas_diario = float(gasto_gas.monto_mensual) / 30 if gasto_gas else 0

    gasto_internet = gastos_servicios_mes.filter(tipo='internet').first()
    gasto_internet_diario = float(
        gasto_internet.monto_mensual) / 30 if gasto_internet else 0

    # Obtener sueldos
    empleados = Usuario.objects.filter(rol__in=['admin', 'operador'])
    empleados_sueldos = []
    total_sueldos_diario = Decimal('0')

    for emp in empleados:
        salario = SalarioEmpleado.objects.filter(empleado=emp).first()
        if salario and salario.salario_semanal > 0:
            sueldo_diario = float(salario.salario_semanal) / 7
            empleados_sueldos.append({
                'nombre': emp.username,
                'rol': 'Admin' if emp.rol == 'admin' else 'Encargado',
                'sueldo_diario': sueldo_diario
            })
            total_sueldos_diario += Decimal(str(sueldo_diario))

    gasto_sueldos_diario = float(total_sueldos_diario)

    # Total gastos del dia
    total_gastos_dia = (gasto_renta_diario + gasto_agua_diario + gasto_luz_diario +
                        gasto_gas_diario + gasto_internet_diario + gasto_sueldos_diario)

    # Utilidad del dia
    utilidad_dia = float(total_ventas) - total_gastos_dia

    context = {
        'fecha': hoy.strftime('%d/%m/%Y'),
        'fecha_hora': timezone.now().strftime('%d/%m/%Y %H:%M'),
        'ventas_efectivo': ventas_efectivo,
        'ventas_tarjeta': ventas_tarjeta,
        'ventas_transferencia': ventas_transferencia,
        'total_ventas': total_ventas,
        'efectivo_contado': efectivo_contado,
        'tarjeta_terminal': tarjeta_terminal,
        'transferencia_banco': transferencia_banco,
        'total_fisico': total_fisico,
        'diferencia': diferencia,
        'justificacion': justificacion,
        'responsable': request.user.username,
        'corte_guardado': corte_existente is not None,

        # Gastos del dia
        'gasto_renta_diario': gasto_renta_diario,
        'gasto_agua_diario': gasto_agua_diario,
        'gasto_luz_diario': gasto_luz_diario,
        'gasto_gas_diario': gasto_gas_diario,
        'gasto_internet_diario': gasto_internet_diario,
        'gasto_sueldos_diario': gasto_sueldos_diario,
        'total_gastos_dia': total_gastos_dia,
        'utilidad_dia': utilidad_dia,
        'empleados_sueldos': empleados_sueldos,
    }

    return render(request, 'admin/finanzas/corte_caja.html', context)

@solo_admin
def admin_historialVentas(request):
    ventas = Pedido.objects.select_related(
        'cliente', 'servicio').order_by('-fecha_recepcion')

    busqueda = request.GET.get('buscar', '').strip()
    if busqueda:
        ventas = ventas.filter(
            Q(folio__icontains=busqueda) |
            Q(cliente__username__icontains=busqueda) |
            Q(cliente__first_name__icontains=busqueda)
        )

    context = {
        'ventas': ventas,
        'total_ventas': ventas.count(),
        'busqueda': busqueda
    }
    return render(request, 'admin/historial/historial-ventas.html', context)


@solo_admin
def admin_historialMovimientos(request):
    movimientos = MovimientoOperador.objects.select_related(
        'operador', 'pedido'
    ).order_by('-fecha')

    operadores = Usuario.objects.filter(
        rol__in=['operador', 'admin']).order_by('username')

    context = {
        'movimientos': movimientos,
        'total_movimientos': movimientos.count(),
        'operadores': operadores
    }
    return render(request, 'admin/historial/historial-movimientos.html', context)

@solo_admin
def admin_detalleVenta(request, pedido_id=None):
    pedido = None
    if pedido_id:
        pedido = get_object_or_404(Pedido, id=pedido_id)

    context = {
        'pedido': pedido
    }
    return render(request, 'admin/historial/detalle-venta.html', context)

@solo_admin
def guardar_gasto_renta(request):
    """Guarda o actualiza el gasto de renta mensual"""
    if request.method == 'POST':
        monto = request.POST.get('monto_renta', 0)
        mes = int(request.POST.get('mes_renta', timezone.now().month))
        anio = int(request.POST.get('anio_renta', timezone.now().year))

        try:
            monto = Decimal(str(monto))

            # Crear o actualizar el registro de renta
            renta, created = GastoRenta.objects.update_or_create(
                mes=mes,
                anio=anio,
                defaults={'monto_mensual': monto}
            )

            if created:
                messages.success(
                    request, f'Renta de {mes}/{anio} registrada exitosamente: ${monto}')
            else:
                messages.success(
                    request, f'Renta de {mes}/{anio} actualizada exitosamente: ${monto}')

        except (ValueError, InvalidOperation) as e:
            messages.error(request, f'Error al guardar la renta: {str(e)}')

    return redirect('finanzas:admin_finanzas')

@solo_admin
def guardar_gasto_servicio(request):
    """Guarda o actualiza el gasto de un servicio (agua, luz, gas, internet)"""
    if request.method == 'POST':
        tipo = request.POST.get('tipo_servicio')
        monto = request.POST.get('monto_servicio', 0)
        mes = int(request.POST.get('mes_servicio', timezone.now().month))
        anio = timezone.now().year

        try:
            monto = Decimal(str(monto))

            # Crear o actualizar el registro del servicio
            servicio, created = GastoServicio.objects.update_or_create(
                tipo=tipo,
                mes=mes,
                anio=anio,
                defaults={'monto_mensual': monto}
            )

            nombre_servicio = dict(
                GastoServicio.TIPOS_SERVICIO).get(tipo, tipo)

            if created:
                messages.success(
                    request, f'{nombre_servicio} de {mes}/{anio} registrado exitosamente: ${monto}')
            else:
                messages.success(
                    request, f'{nombre_servicio} de {mes}/{anio} actualizado exitosamente: ${monto}')

        except (ValueError, InvalidOperation) as e:
            messages.error(request, f'Error al guardar el servicio: {str(e)}')

    return redirect('finanzas:admin_finanzas')

@solo_admin
def guardar_salario_empleado(request):
    """Guarda o actualiza el salario semanal de un empleado"""
    if request.method == 'POST':
        empleado_id = request.POST.get('empleado_id')
        salario = request.POST.get('salario_semanal', 0)

        try:
            salario = Decimal(str(salario))
            empleado = Usuario.objects.get(id=empleado_id)

            # Crear o actualizar el registro del salario
            salario_obj, created = SalarioEmpleado.objects.update_or_create(
                empleado=empleado,
                defaults={'salario_semanal': salario}
            )

            if created:
                messages.success(
                    request, f'Salario de {empleado.username} registrado exitosamente: ${salario}/semana')
            else:
                messages.success(
                    request, f'Salario de {empleado.username} actualizado exitosamente: ${salario}/semana')

        except Usuario.DoesNotExist:
            messages.error(request, 'Empleado no encontrado')
        except (ValueError, InvalidOperation) as e:
            messages.error(request, f'Error al guardar el salario: {str(e)}')

    return redirect('finanzas:admin_finanzas')

@solo_admin
def exportar_finanzas_excel(request):
    from datetime import datetime, timedelta
    from django.db.models import Sum, Count
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    hoy = timezone.now().date()

    filtro = request.GET.get('filtro', 'hoy')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    tab = request.GET.get('tab', 'resumen')

    if filtro == 'hoy':
        fecha_inicio = hoy
        fecha_fin = hoy
        periodo_nombre = f"Hoy - {hoy.strftime('%d/%m/%Y')}"
    elif filtro == 'semana':
        fecha_inicio = hoy - timedelta(days=7)
        fecha_fin = hoy
        periodo_nombre = "Ultima Semana"
    elif filtro == 'mes':
        fecha_inicio = hoy.replace(day=1)
        fecha_fin = hoy
        periodo_nombre = f"Este Mes - {hoy.strftime('%B %Y')}"
    elif filtro == 'personalizado' and fecha_desde and fecha_hasta:
        fecha_inicio = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        periodo_nombre = f"Del {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
    else:
        fecha_inicio = hoy
        fecha_fin = hoy
        periodo_nombre = f"Hoy - {hoy.strftime('%d/%m/%Y')}"

    # Calcular dias del periodo
    dias_periodo = (fecha_fin - fecha_inicio).days + 1

    pedidos_periodo = Pedido.objects.filter(
        fecha_recepcion__date__gte=fecha_inicio,
        fecha_recepcion__date__lte=fecha_fin,
        estado_pago='pagado'
    )

    ingresos_totales = pedidos_periodo.aggregate(
        total=Sum('total'))['total'] or Decimal('0')

    pago_efectivo = pedidos_periodo.filter(metodo_pago='efectivo').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    pago_tarjeta = pedidos_periodo.filter(metodo_pago='tarjeta').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    pago_transferencia = pedidos_periodo.filter(metodo_pago='transferencia').aggregate(
        total=Sum('total'))['total'] or Decimal('0')

    # ========== CALCULAR GASTOS DEL PERIODO ==========
    mes_actual = hoy.month
    anio_actual = hoy.year

    # Obtener renta del mes
    renta_actual = GastoRenta.objects.filter(
        mes=mes_actual, anio=anio_actual).first()
    total_renta_mes = float(renta_actual.monto_mensual) if renta_actual else 0

    # Obtener servicios del mes
    gastos_servicios_mes = GastoServicio.objects.filter(
        mes=mes_actual, anio=anio_actual)
    total_servicios_mes = Decimal('0')
    gastos_servicios_data = []

    tipos_servicio_lista = [
        {'tipo': 'agua', 'nombre': 'Agua'},
        {'tipo': 'luz', 'nombre': 'Luz'},
        {'tipo': 'gas', 'nombre': 'Gas'},
        {'tipo': 'internet', 'nombre': 'Internet'},
    ]

    for ts in tipos_servicio_lista:
        gasto = gastos_servicios_mes.filter(tipo=ts['tipo']).first()
        monto = float(gasto.monto_mensual) if gasto else 0
        total_servicios_mes += Decimal(str(monto))
        if monto > 0:
            gastos_servicios_data.append({
                'tipo': ts['nombre'],
                'monto': monto
            })

    # Obtener sueldos
    empleados = Usuario.objects.filter(rol__in=['admin', 'operador'])
    total_sueldos_semanal = Decimal('0')
    sueldos_data = []

    for emp in empleados:
        salario = SalarioEmpleado.objects.filter(empleado=emp).first()
        salario_semanal = float(salario.salario_semanal) if salario else 0
        salario_diario = salario_semanal / 7 if salario_semanal > 0 else 0

        if salario_semanal > 0:
            sueldos_data.append({
                'nombre': f"{emp.first_name} {emp.last_name}".strip() or emp.username,
                'rol': 'Administrador' if emp.rol == 'admin' else 'Encargado',
                'sueldo_semanal': salario_semanal,
                'sueldo_diario': salario_diario
            })
            total_sueldos_semanal += Decimal(str(salario_semanal))

    total_sueldos_mes = float(total_sueldos_semanal * 4)

    # Gastos proporcionales al periodo
    gasto_renta_diario = total_renta_mes / 30 if total_renta_mes > 0 else 0
    gasto_servicios_diario = float(total_servicios_mes) / 30 if total_servicios_mes > 0 else 0
    gasto_sueldos_diario = float(total_sueldos_semanal) / 7 if total_sueldos_semanal > 0 else 0

    gastos_renta_periodo = gasto_renta_diario * dias_periodo
    gastos_servicios_periodo = gasto_servicios_diario * dias_periodo
    gastos_sueldos_periodo = gasto_sueldos_diario * dias_periodo
    total_gastos_periodo = gastos_renta_periodo + gastos_servicios_periodo + gastos_sueldos_periodo

    # Utilidad neta
    utilidad_neta = float(ingresos_totales) - total_gastos_periodo

    detalles_periodo = DetallePedido.objects.filter(
        pedido__fecha_recepcion__date__gte=fecha_inicio,
        pedido__fecha_recepcion__date__lte=fecha_fin,
        pedido__estado_pago='pagado'
    )
    prendas_stats = detalles_periodo.values(
        'prenda__nombre'
    ).annotate(
        cantidad_total=Sum('cantidad'),
        ganancia_total=Sum('subtotal')
    ).order_by('-cantidad_total')

    servicios_stats = pedidos_periodo.values(
        'tipo_servicio'
    ).annotate(
        cantidad=Count('id'),
        ganancia_total=Sum('total')
    ).order_by('-cantidad')

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Financiero"

    header_fill = PatternFill(start_color="2d3748",
                              end_color="2d3748", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=14)
    subtotal_fill = PatternFill(
        start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")
    gastos_fill = PatternFill(
        start_color="fee2e2", end_color="fee2e2", fill_type="solid")
    positive_font = Font(color="28a745", bold=True)
    negative_font = Font(color="dc3545", bold=True)

    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = "REPORTE FINANCIERO - PUNTO LIMPIO"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:D2')
    cell = ws['A2']
    cell.value = periodo_nombre
    cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 4

    # ========== SECCION RESUMEN (siempre incluir o si tab es resumen) ==========
    if tab in ['resumen', 'ingresos']:
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "RESUMEN FINANCIERO"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

        row += 1
        ws[f'A{row}'] = "Concepto"
        ws[f'B{row}'] = "Monto"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Ingresos totales"
        ws[f'B{row}'] = float(ingresos_totales)
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = positive_font

        row += 1
        ws[f'A{row}'] = "Gastos totales"
        ws[f'B{row}'] = float(total_gastos_periodo)
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = negative_font

        row += 1
        ws[f'A{row}'] = "UTILIDAD NETA"
        ws[f'B{row}'] = float(utilidad_neta)
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True, color="28a745" if utilidad_neta >= 0 else "dc3545")
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'A{row}'].fill = subtotal_fill
        ws[f'B{row}'].fill = subtotal_fill

        row += 2

    # ========== SECCION INGRESOS ==========
    if tab in ['resumen', 'ingresos']:
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "DESGLOSE POR METODO DE PAGO"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

        row += 1
        ws[f'A{row}'] = "Metodo"
        ws[f'B{row}'] = "Monto"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Efectivo"
        ws[f'B{row}'] = float(pago_efectivo)
        ws[f'B{row}'].number_format = '$#,##0.00'

        row += 1
        ws[f'A{row}'] = "Tarjeta"
        ws[f'B{row}'] = float(pago_tarjeta)
        ws[f'B{row}'].number_format = '$#,##0.00'

        row += 1
        ws[f'A{row}'] = "Transferencia"
        ws[f'B{row}'] = float(pago_transferencia)
        ws[f'B{row}'].number_format = '$#,##0.00'

        row += 2

        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "ESTADISTICAS POR PRENDA"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

        row += 1
        ws[f'A{row}'] = "Prenda"
        ws[f'B{row}'] = "Cantidad"
        ws[f'C{row}'] = "Ganancia"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)

        for prenda in prendas_stats:
            if prenda['prenda__nombre']:
                row += 1
                ws[f'A{row}'] = prenda['prenda__nombre']
                ws[f'B{row}'] = prenda['cantidad_total']
                ws[f'C{row}'] = float(prenda['ganancia_total'] or 0)
                ws[f'C{row}'].number_format = '$#,##0.00'

        row += 2
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "ESTADISTICAS POR SERVICIO"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

        row += 1
        ws[f'A{row}'] = "Servicio"
        ws[f'B{row}'] = "Cantidad"
        ws[f'C{row}'] = "Ganancia"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)

        for servicio in servicios_stats:
            row += 1
            ws[f'A{row}'] = servicio['tipo_servicio'] or 'Sin especificar'
            ws[f'B{row}'] = servicio['cantidad']
            ws[f'C{row}'] = float(servicio['ganancia_total'] or 0)
            ws[f'C{row}'].number_format = '$#,##0.00'

        row += 2

    # ========== SECCION GASTOS ==========
    if tab in ['resumen', 'gastos']:
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "GASTOS DEL PERIODO"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

        row += 1
        ws[f'A{row}'] = "Concepto"
        ws[f'B{row}'] = "Monto"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Renta (proporcional)"
        ws[f'B{row}'] = float(gastos_renta_periodo)
        ws[f'B{row}'].number_format = '$#,##0.00'

        row += 1
        ws[f'A{row}'] = "Servicios (proporcional)"
        ws[f'B{row}'] = float(gastos_servicios_periodo)
        ws[f'B{row}'].number_format = '$#,##0.00'

        row += 1
        ws[f'A{row}'] = "Sueldos (proporcional)"
        ws[f'B{row}'] = float(gastos_sueldos_periodo)
        ws[f'B{row}'].number_format = '$#,##0.00'

        row += 1
        ws[f'A{row}'] = "TOTAL GASTOS"
        ws[f'B{row}'] = float(total_gastos_periodo)
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = negative_font
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'A{row}'].fill = gastos_fill
        ws[f'B{row}'].fill = gastos_fill

        row += 2

        # Detalle de servicios
        if gastos_servicios_data:
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = "DETALLE DE GASTOS POR SERVICIOS"
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

            row += 1
            ws[f'A{row}'] = "Servicio"
            ws[f'B{row}'] = "Monto Mensual"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)

            for gasto in gastos_servicios_data:
                row += 1
                ws[f'A{row}'] = gasto['tipo']
                ws[f'B{row}'] = gasto['monto']
                ws[f'B{row}'].number_format = '$#,##0.00'

            row += 2

        # Detalle de sueldos
        if sueldos_data:
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = "DETALLE DE SUELDOS"
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

            row += 1
            ws[f'A{row}'] = "Empleado"
            ws[f'B{row}'] = "Rol"
            ws[f'C{row}'] = "Sueldo Semanal"
            ws[f'D{row}'] = "Sueldo Diario"
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].font = Font(bold=True)

            for sueldo in sueldos_data:
                row += 1
                ws[f'A{row}'] = sueldo['nombre']
                ws[f'B{row}'] = sueldo['rol']
                ws[f'C{row}'] = sueldo['sueldo_semanal']
                ws[f'C{row}'].number_format = '$#,##0.00'
                ws[f'D{row}'] = sueldo['sueldo_diario']
                ws[f'D{row}'].number_format = '$#,##0.00'

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"reporte_financiero_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

@solo_admin
def exportar_historial_ventas_excel(request):
    """Exportar historial de ventas a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # Obtener parametros de busqueda
    search = request.GET.get('search', '')

    # Obtener ventas
    ventas = Pedido.objects.select_related('cliente').order_by('-fecha_recepcion')

    if search:
        ventas = ventas.filter(
            Q(folio__icontains=search) |
            Q(cliente__first_name__icontains=search) |
            Q(cliente__last_name__icontains=search) |
            Q(cliente__username__icontains=search)
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de Ventas"

    # Estilos
    header_fill = PatternFill(start_color="2d3748", end_color="2d3748", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Titulo
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = "HISTORIAL DE VENTAS - PUNTO LIMPIO"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:G2')
    cell = ws['A2']
    cell.value = f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    cell.alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Folio', 'Cliente', 'Tipo de Servicio', 'Total', 'Estado de Pago', 'Metodo de Pago', 'Fecha']
    row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Datos
    for venta in ventas:
        row += 1
        cliente_nombre = f"{venta.cliente.first_name} {venta.cliente.last_name}".strip()
        if not cliente_nombre:
            cliente_nombre = venta.cliente.username

        estado_pago = "Pagado" if venta.estado_pago == 'pagado' else "Pendiente"
        metodo_pago = venta.metodo_pago.capitalize() if venta.metodo_pago else "No especificado"

        data = [
            venta.folio,
            cliente_nombre,
            venta.tipo_servicio or 'Sin especificar',
            float(venta.total),
            estado_pago,
            metodo_pago,
            venta.fecha_recepcion.strftime('%d/%m/%Y %H:%M')
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col == 4:  # Total
                cell.number_format = '$#,##0.00'

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18

    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"historial_ventas_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

@solo_admin
def exportar_historial_movimientos_excel(request):
    """Exportar historial de movimientos a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # Obtener parametros de filtro
    usuario = request.GET.get('usuario', '')
    accion = request.GET.get('accion', '')
    fecha = request.GET.get('fecha', '')

    # Obtener movimientos
    movimientos = MovimientoOperador.objects.select_related('operador').order_by('-fecha')

    if usuario:
        movimientos = movimientos.filter(operador__username__icontains=usuario)

    if accion:
        movimientos = movimientos.filter(accion__icontains=accion)

    if fecha:
        hoy = timezone.now().date()
        if fecha == '24h':
            fecha_inicio = hoy - timedelta(days=1)
            movimientos = movimientos.filter(fecha__date__gte=fecha_inicio)
        elif fecha == '7d':
            fecha_inicio = hoy - timedelta(days=7)
            movimientos = movimientos.filter(fecha__date__gte=fecha_inicio)
        elif fecha == '30d':
            fecha_inicio = hoy - timedelta(days=30)
            movimientos = movimientos.filter(fecha__date__gte=fecha_inicio)

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de Movimientos"

    # Estilos
    header_fill = PatternFill(start_color="2d3748", end_color="2d3748", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Titulo
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = "HISTORIAL DE MOVIMIENTOS - PUNTO LIMPIO"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:D2')
    cell = ws['A2']
    cell.value = f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    cell.alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Fecha/Hora', 'Usuario', 'Accion', 'Detalles']
    row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Diccionario de acciones
    acciones_display = {
        'registro_servicio': 'Registro servicio',
        'entrego': 'Entrego',
        'cambio_precio': 'Cambio precio',
        'creo_ticket': 'Creo ticket',
        'elimino': 'Elimino',
        'actualizo': 'Actualizo',
    }

    # Datos
    for movimiento in movimientos:
        row += 1
        accion_display = acciones_display.get(movimiento.accion, movimiento.accion)

        data = [
            movimiento.fecha.strftime('%d/%m/%Y %H:%M'),
            movimiento.operador.username if movimiento.operador else 'Sistema',
            accion_display,
            movimiento.detalles or ''
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 50

    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"historial_movimientos_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

@solo_admin
def imprimir_reporte_finanzas(request):
    from datetime import datetime, timedelta
    from django.db.models import Sum, Count
    from django.template.loader import get_template
    from xhtml2pdf import pisa
    from io import BytesIO

    hoy = timezone.now().date()

    filtro = request.GET.get('filtro', 'hoy')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    tab = request.GET.get('tab', 'resumen')

    if filtro == 'hoy':
        fecha_inicio = hoy
        fecha_fin = hoy
    elif filtro == 'semana':
        fecha_inicio = hoy - timedelta(days=7)
        fecha_fin = hoy
    elif filtro == 'mes':
        fecha_inicio = hoy.replace(day=1)
        fecha_fin = hoy
    elif filtro == 'personalizado' and fecha_desde and fecha_hasta:
        fecha_inicio = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    else:
        fecha_inicio = hoy
        fecha_fin = hoy

    # Calcular dias del periodo
    dias_periodo = (fecha_fin - fecha_inicio).days + 1

    pedidos_periodo = Pedido.objects.filter(
        fecha_recepcion__date__gte=fecha_inicio,
        fecha_recepcion__date__lte=fecha_fin,
        estado_pago='pagado'
    )

    ingresos_totales = pedidos_periodo.aggregate(
        total=Sum('total'))['total'] or Decimal('0')

    pago_efectivo = pedidos_periodo.filter(metodo_pago='efectivo').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    pago_tarjeta = pedidos_periodo.filter(metodo_pago='tarjeta').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    pago_transferencia = pedidos_periodo.filter(metodo_pago='transferencia').aggregate(
        total=Sum('total'))['total'] or Decimal('0')

    total_pagos = pago_efectivo + pago_tarjeta + pago_transferencia
    pct_efectivo = round((pago_efectivo / total_pagos * 100),
                         1) if total_pagos > 0 else 0
    pct_tarjeta = round((pago_tarjeta / total_pagos * 100),
                        1) if total_pagos > 0 else 0
    pct_transferencia = round(
        (pago_transferencia / total_pagos * 100), 1) if total_pagos > 0 else 0

    # ========== CALCULAR GASTOS DEL PERIODO ==========
    mes_actual = hoy.month
    anio_actual = hoy.year

    # Obtener renta del mes
    renta_actual = GastoRenta.objects.filter(
        mes=mes_actual, anio=anio_actual).first()
    total_renta_mes = float(renta_actual.monto_mensual) if renta_actual else 0

    # Obtener servicios del mes
    gastos_servicios_mes = GastoServicio.objects.filter(
        mes=mes_actual, anio=anio_actual)
    total_servicios_mes = Decimal('0')
    gastos_servicios_data = []

    tipos_servicio_lista = [
        {'tipo': 'agua', 'nombre': 'Agua'},
        {'tipo': 'luz', 'nombre': 'Luz'},
        {'tipo': 'gas', 'nombre': 'Gas'},
        {'tipo': 'internet', 'nombre': 'Internet'},
    ]

    for ts in tipos_servicio_lista:
        gasto = gastos_servicios_mes.filter(tipo=ts['tipo']).first()
        monto = float(gasto.monto_mensual) if gasto else 0
        total_servicios_mes += Decimal(str(monto))
        if monto > 0:
            gastos_servicios_data.append({
                'tipo': ts['nombre'],
                'monto': monto
            })

    # Obtener sueldos
    empleados = Usuario.objects.filter(rol__in=['admin', 'operador'])
    total_sueldos_semanal = Decimal('0')
    sueldos_data = []

    for emp in empleados:
        salario = SalarioEmpleado.objects.filter(empleado=emp).first()
        salario_semanal = float(salario.salario_semanal) if salario else 0
        salario_diario = salario_semanal / 7 if salario_semanal > 0 else 0

        if salario_semanal > 0:
            sueldos_data.append({
                'nombre': f"{emp.first_name} {emp.last_name}".strip() or emp.username,
                'rol': 'Administrador' if emp.rol == 'admin' else 'Encargado',
                'sueldo_semanal': salario_semanal,
                'sueldo_diario': salario_diario
            })
            total_sueldos_semanal += Decimal(str(salario_semanal))

    total_sueldos_mes = float(total_sueldos_semanal * 4)

    # Gastos proporcionales al periodo
    gasto_renta_diario = total_renta_mes / 30 if total_renta_mes > 0 else 0
    gasto_servicios_diario = float(total_servicios_mes) / 30 if total_servicios_mes > 0 else 0
    gasto_sueldos_diario = float(total_sueldos_semanal) / 7 if total_sueldos_semanal > 0 else 0

    gastos_renta_periodo = gasto_renta_diario * dias_periodo
    gastos_servicios_periodo = gasto_servicios_diario * dias_periodo
    gastos_sueldos_periodo = gasto_sueldos_diario * dias_periodo
    total_gastos_periodo = gastos_renta_periodo + gastos_servicios_periodo + gastos_sueldos_periodo

    # Utilidad neta
    utilidad_neta = float(ingresos_totales) - total_gastos_periodo

    detalles_periodo = DetallePedido.objects.filter(
        pedido__fecha_recepcion__date__gte=fecha_inicio,
        pedido__fecha_recepcion__date__lte=fecha_fin,
        pedido__estado_pago='pagado'
    )
    prendas_stats = detalles_periodo.values(
        'prenda__nombre'
    ).annotate(
        cantidad_total=Sum('cantidad'),
        ganancia_total=Sum('subtotal')
    ).order_by('-cantidad_total')[:10]

    servicios_stats = pedidos_periodo.values(
        'tipo_servicio'
    ).annotate(
        cantidad=Count('id'),
        ganancia_total=Sum('total')
    ).order_by('-cantidad')

    context = {
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'dias_periodo': dias_periodo,
        'tab': tab,
        'ingresos_totales': ingresos_totales,
        'utilidad_neta': utilidad_neta,
        'pago_efectivo': pago_efectivo,
        'pago_tarjeta': pago_tarjeta,
        'pago_transferencia': pago_transferencia,
        'pct_efectivo': pct_efectivo,
        'pct_tarjeta': pct_tarjeta,
        'pct_transferencia': pct_transferencia,
        'prendas_stats': prendas_stats,
        'servicios_stats': servicios_stats,
        # Gastos
        'gastos_renta_periodo': gastos_renta_periodo,
        'gastos_servicios_periodo': gastos_servicios_periodo,
        'gastos_sueldos_periodo': gastos_sueldos_periodo,
        'total_gastos_periodo': total_gastos_periodo,
        'gastos_servicios_data': gastos_servicios_data,
        'sueldos_data': sueldos_data,
        'total_renta_mes': total_renta_mes,
        'total_servicios_mes': float(total_servicios_mes),
        'total_sueldos_mes': total_sueldos_mes,
    }

    template = get_template('admin/finanzas/reporte_finanzas_pdf.html')
    html = template.render(context)

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if pdf.err:
        return HttpResponse("Error al generar el PDF", status=500)

    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    filename = f"reporte_financiero_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response

@solo_admin
@require_POST
def enviar_reporte_email(request):
    """Vista para enviar el reporte financiero por correo electronico"""
    if not request.user.groups.filter(name='Administrador').exists():
        return JsonResponse({'success': False, 'message': 'No autorizado'}, status=403)

    try:
        # Obtener los datos del request
        data = json.loads(request.body)
        email_destino = data.get('email')
        filtro = data.get('filtro', 'hoy')
        fecha_desde = data.get('fecha_desde')
        fecha_hasta = data.get('fecha_hasta')
        tab = data.get('tab', 'resumen')

        if not email_destino:
            return JsonResponse({'success': False, 'message': 'Email requerido'}, status=400)

        # Validar formato de email
        from django.core.validators import validate_email
        from django.core.exceptions import ValidationError
        try:
            validate_email(email_destino)
        except ValidationError:
            return JsonResponse({'success': False, 'message': 'Email invalido'}, status=400)
        
        # Calcular fechas segun el filtro
        hoy = timezone.now().date()
        if filtro == 'hoy':
            fecha_inicio = hoy
            fecha_fin = hoy
        elif filtro == 'semana':
            fecha_inicio = hoy - timedelta(days=7)
            fecha_fin = hoy
        elif filtro == 'mes':
            fecha_inicio = hoy.replace(day=1)
            fecha_fin = hoy
        elif filtro == 'personalizado' and fecha_desde and fecha_hasta:
            fecha_inicio = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        else:
            fecha_inicio = hoy
            fecha_fin = hoy

        # Calcular dias del periodo
        dias_periodo = (fecha_fin - fecha_inicio).days + 1

        # Obtener datos financieros
        pedidos_periodo = Pedido.objects.filter(
            fecha_recepcion__date__gte=fecha_inicio,
            fecha_recepcion__date__lte=fecha_fin,
            estado_pago='pagado'
        )

        ingresos_totales = pedidos_periodo.aggregate(
            total=Sum('total'))['total'] or Decimal('0')

        # Metodos de pago
        pago_efectivo = pedidos_periodo.filter(metodo_pago='efectivo').aggregate(
            total=Sum('total'))['total'] or Decimal('0')
        pago_tarjeta = pedidos_periodo.filter(metodo_pago='tarjeta').aggregate(
            total=Sum('total'))['total'] or Decimal('0')
        pago_transferencia = pedidos_periodo.filter(metodo_pago='transferencia').aggregate(
            total=Sum('total'))['total'] or Decimal('0')

        total_pagos = pago_efectivo + pago_tarjeta + pago_transferencia
        pct_efectivo = round(
            (pago_efectivo / total_pagos * 100), 1) if total_pagos > 0 else 0
        pct_tarjeta = round((pago_tarjeta / total_pagos * 100),
                            1) if total_pagos > 0 else 0
        pct_transferencia = round(
            (pago_transferencia / total_pagos * 100), 1) if total_pagos > 0 else 0

        # ========== CALCULAR GASTOS DEL PERIODO ==========
        mes_actual = hoy.month
        anio_actual = hoy.year

        # Obtener renta del mes
        renta_actual = GastoRenta.objects.filter(
            mes=mes_actual, anio=anio_actual).first()
        total_renta_mes = float(renta_actual.monto_mensual) if renta_actual else 0

        # Obtener servicios del mes
        gastos_servicios_mes = GastoServicio.objects.filter(
            mes=mes_actual, anio=anio_actual)
        total_servicios_mes = Decimal('0')
        gastos_servicios_data = []

        tipos_servicio_lista = [
            {'tipo': 'agua', 'nombre': 'Agua'},
            {'tipo': 'luz', 'nombre': 'Luz'},
            {'tipo': 'gas', 'nombre': 'Gas'},
            {'tipo': 'internet', 'nombre': 'Internet'},
        ]

        for ts in tipos_servicio_lista:
            gasto = gastos_servicios_mes.filter(tipo=ts['tipo']).first()
            monto = float(gasto.monto_mensual) if gasto else 0
            total_servicios_mes += Decimal(str(monto))
            if monto > 0:
                gastos_servicios_data.append({
                    'tipo': ts['nombre'],
                    'monto': monto
                })

        # Obtener sueldos
        empleados = Usuario.objects.filter(rol__in=['admin', 'operador'])
        total_sueldos_semanal = Decimal('0')
        sueldos_data = []

        for emp in empleados:
            salario = SalarioEmpleado.objects.filter(empleado=emp).first()
            salario_semanal = float(salario.salario_semanal) if salario else 0
            salario_diario = salario_semanal / 7 if salario_semanal > 0 else 0

            if salario_semanal > 0:
                sueldos_data.append({
                    'nombre': f"{emp.first_name} {emp.last_name}".strip() or emp.username,
                    'rol': 'Administrador' if emp.rol == 'admin' else 'Encargado',
                    'sueldo_semanal': salario_semanal,
                    'sueldo_diario': salario_diario
                })
                total_sueldos_semanal += Decimal(str(salario_semanal))

        total_sueldos_mes = float(total_sueldos_semanal * 4)

        # Gastos proporcionales al periodo
        gasto_renta_diario = total_renta_mes / 30 if total_renta_mes > 0 else 0
        gasto_servicios_diario = float(total_servicios_mes) / 30 if total_servicios_mes > 0 else 0
        gasto_sueldos_diario = float(total_sueldos_semanal) / 7 if total_sueldos_semanal > 0 else 0

        gastos_renta_periodo = gasto_renta_diario * dias_periodo
        gastos_servicios_periodo = gasto_servicios_diario * dias_periodo
        gastos_sueldos_periodo = gasto_sueldos_diario * dias_periodo
        total_gastos_periodo = gastos_renta_periodo + gastos_servicios_periodo + gastos_sueldos_periodo

        # Utilidad neta
        utilidad_neta = float(ingresos_totales) - total_gastos_periodo

        # Datos de prendas
        detalles_periodo = DetallePedido.objects.filter(
            pedido__fecha_recepcion__date__gte=fecha_inicio,
            pedido__fecha_recepcion__date__lte=fecha_fin,
            pedido__estado_pago='pagado'
        )
        prendas_stats = detalles_periodo.values(
            'prenda__nombre'
        ).annotate(
            cantidad_total=Sum('cantidad'),
            ganancia_total=Sum('subtotal')
        ).order_by('-cantidad_total')[:10]

        # Datos de servicios
        servicios_stats = pedidos_periodo.values(
            'tipo_servicio'
        ).annotate(
            cantidad=Count('id'),
            ganancia_total=Sum('total')
        ).order_by('-cantidad')

        # Preparar contexto para el PDF
        context = {
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'dias_periodo': dias_periodo,
            'tab': tab,
            'ingresos_totales': ingresos_totales,
            'utilidad_neta': utilidad_neta,
            'pago_efectivo': pago_efectivo,
            'pago_tarjeta': pago_tarjeta,
            'pago_transferencia': pago_transferencia,
            'pct_efectivo': pct_efectivo,
            'pct_tarjeta': pct_tarjeta,
            'pct_transferencia': pct_transferencia,
            'prendas_stats': prendas_stats,
            'servicios_stats': servicios_stats,
            # Gastos
            'gastos_renta_periodo': gastos_renta_periodo,
            'gastos_servicios_periodo': gastos_servicios_periodo,
            'gastos_sueldos_periodo': gastos_sueldos_periodo,
            'total_gastos_periodo': total_gastos_periodo,
            'gastos_servicios_data': gastos_servicios_data,
            'sueldos_data': sueldos_data,
            'total_renta_mes': total_renta_mes,
            'total_servicios_mes': float(total_servicios_mes),
            'total_sueldos_mes': total_sueldos_mes,
        }

        # Generar PDF
        from django.template.loader import get_template
        from io import BytesIO
        from xhtml2pdf import pisa

        template = get_template('admin/finanzas/reporte_finanzas_pdf.html')
        html = template.render(context)

        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

        if pdf.err:
            return JsonResponse({'success': False, 'message': 'Error al generar el PDF'}, status=500)

        pdf_bytes = result.getvalue()

        # Enviar email
        from django.core.mail import EmailMessage
        from django.conf import settings

        # Determinar el nombre del periodo
        if filtro == 'hoy':
            periodo_nombre = f"del dia {fecha_inicio.strftime('%d/%m/%Y')}"
        elif filtro == 'semana':
            periodo_nombre = "de la ultima semana"
        elif filtro == 'mes':
            periodo_nombre = "del mes actual"
        else:
            periodo_nombre = f"del {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"

        subject = f'Reporte Financiero Punto Limpio - {periodo_nombre}'
        body = f'''Hola,

Adjunto encontraras el reporte financiero de Punto Limpio {periodo_nombre}.

RESUMEN DEL PERIODO:
- Ingresos totales: ${float(ingresos_totales):,.2f}
- Gastos totales: ${total_gastos_periodo:,.2f}
- Utilidad neta: ${utilidad_neta:,.2f}

DESGLOSE DE GASTOS:
- Renta (proporcional): ${gastos_renta_periodo:,.2f}
- Servicios (proporcional): ${gastos_servicios_periodo:,.2f}
- Sueldos (proporcional): ${gastos_sueldos_periodo:,.2f}

METODOS DE PAGO:
- Efectivo: ${float(pago_efectivo):,.2f} ({pct_efectivo}%)
- Tarjeta: ${float(pago_tarjeta):,.2f} ({pct_tarjeta}%)
- Transferencia: ${float(pago_transferencia):,.2f} ({pct_transferencia}%)

Este reporte fue generado automaticamente por {request.user.username} el {timezone.now().strftime('%d/%m/%Y a las %H:%M')}.

Saludos,
Sistema Punto Limpio
'''

        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=settings.EMAIL_HOST_USER,
            to=[email_destino],
        )

        # Adjuntar el PDF
        filename = f"reporte_financiero_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.pdf"
        email.attach(filename, pdf_bytes, 'application/pdf')

        # Enviar el email
        email.send()

        return JsonResponse({
            'success': True,
            'message': f'Reporte enviado exitosamente a {email_destino}'
        })

    except Exception as e:
        print(f"Error enviando reporte por email: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error al enviar el reporte: {str(e)}'
        }, status=500)

@solo_admin
def imprimir_corte_caja(request):
    """
    Genera un PDF del corte de caja del día
    """
    if not request.user.groups.filter(name='Administrador').exists():
        return HttpResponse("No autorizado", status=403)

    from django.template.loader import get_template
    from xhtml2pdf import pisa
    from io import BytesIO

    # Obtener fecha de hoy
    hoy = timezone.now().date()

    # Pedidos pagados del día de hoy
    pedidos_hoy = Pedido.objects.filter(
        fecha_recepcion__date=hoy,
        estado_pago='pagado'
    )

    # Calcular ventas por método de pago
    ventas_efectivo = pedidos_hoy.filter(metodo_pago='efectivo').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    ventas_tarjeta = pedidos_hoy.filter(metodo_pago='tarjeta').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    ventas_transferencia = pedidos_hoy.filter(metodo_pago='transferencia').aggregate(
        total=Sum('total'))['total'] or Decimal('0')

    total_ventas = ventas_efectivo + ventas_tarjeta + ventas_transferencia

    # Obtener corte guardado si existe
    corte_existente = CorteCaja.objects.filter(
        fecha=hoy, responsable=request.user).first()

    if corte_existente:
        efectivo_contado = corte_existente.efectivo_contado
        tarjeta_terminal = corte_existente.tarjeta_terminal
        transferencia_banco = corte_existente.transferencia_banco
        total_fisico = corte_existente.total_fisico
        diferencia = corte_existente.diferencia
        justificacion = corte_existente.justificacion or ''
    else:
        efectivo_contado = Decimal('0')
        tarjeta_terminal = Decimal('0')
        transferencia_banco = Decimal('0')
        total_fisico = Decimal('0')
        diferencia = Decimal('0')
        justificacion = ''

    # ========== CALCULAR GASTOS DEL DIA ==========
    mes_actual = hoy.month
    anio_actual = hoy.year

    # Obtener renta del mes
    renta_actual = GastoRenta.objects.filter(
        mes=mes_actual, anio=anio_actual).first()
    gasto_renta_diario = float(
        renta_actual.monto_mensual) / 30 if renta_actual else 0

    # Obtener servicios del mes
    gastos_servicios_mes = GastoServicio.objects.filter(
        mes=mes_actual, anio=anio_actual)

    gasto_agua = gastos_servicios_mes.filter(tipo='agua').first()
    gasto_agua_diario = float(
        gasto_agua.monto_mensual) / 30 if gasto_agua else 0

    gasto_luz = gastos_servicios_mes.filter(tipo='luz').first()
    gasto_luz_diario = float(gasto_luz.monto_mensual) / 30 if gasto_luz else 0

    gasto_gas = gastos_servicios_mes.filter(tipo='gas').first()
    gasto_gas_diario = float(gasto_gas.monto_mensual) / 30 if gasto_gas else 0

    gasto_internet = gastos_servicios_mes.filter(tipo='internet').first()
    gasto_internet_diario = float(
        gasto_internet.monto_mensual) / 30 if gasto_internet else 0

    # Obtener sueldos
    empleados_pdf = Usuario.objects.filter(rol__in=['admin', 'operador'])
    total_sueldos_diario_pdf = Decimal('0')

    for emp in empleados_pdf:
        salario = SalarioEmpleado.objects.filter(empleado=emp).first()
        if salario and salario.salario_semanal > 0:
            sueldo_diario = float(salario.salario_semanal) / 7
            total_sueldos_diario_pdf += Decimal(str(sueldo_diario))

    gasto_sueldos_diario = float(total_sueldos_diario_pdf)

    # Total gastos del dia
    total_gastos_dia = (gasto_renta_diario + gasto_agua_diario + gasto_luz_diario +
                        gasto_gas_diario + gasto_internet_diario + gasto_sueldos_diario)

    # Utilidad del dia
    utilidad_dia = float(total_ventas) - total_gastos_dia

    context = {
        'fecha': hoy.strftime('%d/%m/%Y'),
        'fecha_hora': timezone.now().strftime('%d/%m/%Y %H:%M'),
        'ventas_efectivo': ventas_efectivo,
        'ventas_tarjeta': ventas_tarjeta,
        'ventas_transferencia': ventas_transferencia,
        'total_ventas': total_ventas,
        'efectivo_contado': efectivo_contado,
        'tarjeta_terminal': tarjeta_terminal,
        'transferencia_banco': transferencia_banco,
        'total_fisico': total_fisico,
        'diferencia': diferencia,
        'justificacion': justificacion,
        'responsable': request.user.username,
        # Gastos del dia
        'gasto_renta_diario': gasto_renta_diario,
        'gasto_agua_diario': gasto_agua_diario,
        'gasto_luz_diario': gasto_luz_diario,
        'gasto_gas_diario': gasto_gas_diario,
        'gasto_internet_diario': gasto_internet_diario,
        'gasto_sueldos_diario': gasto_sueldos_diario,
        'total_gastos_dia': total_gastos_dia,
        'utilidad_dia': utilidad_dia,
    }

    # Renderizar template
    template_path = 'admin/finanzas/corte_caja_pdf.html'
    template = get_template(template_path)
    html = template.render(context)

    # Crear PDF
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if pdf.err:
        return HttpResponse('Error al generar el PDF', status=500)

    # Retornar PDF
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="corte_caja_{hoy.strftime("%Y%m%d")}.pdf"'

    return response