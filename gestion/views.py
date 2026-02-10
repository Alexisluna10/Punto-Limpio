from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.contrib.auth.models import Group
from django.utils import timezone
from django.db.models import Q, Sum, Count
from decimal import Decimal
from django.contrib.auth.hashers import make_password
import json
from datetime import datetime, timedelta
from django.views.decorators.cache import never_cache
# Importaciones locales
from .utils import render_pdf_ticket, enviar_ticket_email, descontar_insumos_por_pedido, descontar_insumos_por_servicio_autoservicio
from .decorators import solo_cliente, solo_trabajador, solo_admin
from django.urls import reverse

# Modelos
from usuarios.models import Usuario
from usuarios.forms import RegistroUsuarioAdminForm
from .models import (
    Insumo, NotificacionStock, Prenda, Servicio, Pedido,
    DetallePedido, MovimientoOperador, Maquina,
    Incidencia, DudaQueja, CorteCaja, GastoRenta, GastoServicio, SalarioEmpleado
)
from .forms_inventario import InsumoForm


def prueba(request):
    return HttpResponse("Prueba app gestión")


# ==========================================
#              VISTAS ADMINISTRADOR
# ==========================================

@solo_admin
def admin_dashboard(request):
    if not request.user.groups.filter(name='Administrador').exists():
        return redirect('tasks')

    hoy = timezone.now().date()
    inicio_semana = hoy - timedelta(days=7)
    inicio_mes = hoy.replace(day=1)

    # ========== GANANCIAS ==========
    # Hoy
    pedidos_hoy = Pedido.objects.filter(
        fecha_recepcion__date=hoy,
        estado_pago='pagado'
    )
    ganancias_hoy = pedidos_hoy.aggregate(total=Sum('total'))[
        'total'] or Decimal('0')
    servicios_hoy = pedidos_hoy.count()

    # Esta semana
    pedidos_semana = Pedido.objects.filter(
        fecha_recepcion__date__gte=inicio_semana,
        fecha_recepcion__date__lte=hoy,
        estado_pago='pagado'
    )
    ganancias_semana = pedidos_semana.aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    servicios_semana = pedidos_semana.count()

    # Este mes
    pedidos_mes = Pedido.objects.filter(
        fecha_recepcion__date__gte=inicio_mes,
        fecha_recepcion__date__lte=hoy,
        estado_pago='pagado'
    )
    ganancias_mes = pedidos_mes.aggregate(total=Sum('total'))[
        'total'] or Decimal('0')
    servicios_mes = pedidos_mes.count()

    # ========== ALERTAS CRÍTICAS ==========
    # Insumos con stock crítico (<=10%)
    insumos_criticos = Insumo.objects.all()
    alertas_insumos = []
    for insumo in insumos_criticos:
        if insumo.porcentaje() <= 10:
            alertas_insumos.append({
                'texto': f'{insumo.nombre} al {insumo.porcentaje()}% de stock',
                'tipo': 'insumo'
            })

    # Incidencias recientes del personal (últimas 3 pendientes o en proceso)
    incidencias_recientes = Incidencia.objects.filter(
        estado__in=['pendiente', 'en_proceso']
    ).order_by('-fecha_reporte')[:3]

    # Dudas/Quejas recientes de clientes (últimas 3 pendientes o en proceso)
    dudas_recientes = DudaQueja.objects.filter(
        estado__in=['pendiente', 'en_proceso']
    ).order_by('-fecha_creacion')[:3]

    # ========== SERVICIOS ACTIVOS ==========
    servicios_totales = Pedido.objects.exclude(estado='entregado').count()
    servicios_pendientes = Pedido.objects.filter(estado='pendiente').count()
    servicios_proceso = Pedido.objects.filter(estado='en_proceso').count()
    servicios_listos = Pedido.objects.filter(estado='listo').count()

    # Máquinas en uso (lavado/secado)
    maquinas_lavado = Maquina.objects.filter(
        tipo='lavadora', estado='ocupado').count()
    maquinas_secado = Maquina.objects.filter(
        tipo='secadora', estado='ocupado').count()

    # ========== PRECIOS DE PRENDAS ==========
    # Obtener 5 prendas destacadas (las más caras o populares)
    prendas_destacadas = Prenda.objects.filter(
        activo=True).order_by('-precio')[:5]

    context = {
        # Ganancias
        'ganancias_hoy': ganancias_hoy,
        'servicios_hoy': servicios_hoy,
        'ganancias_semana': ganancias_semana,
        'servicios_semana': servicios_semana,
        'ganancias_mes': ganancias_mes,
        'servicios_mes': servicios_mes,

        # Alertas
        'alertas_insumos': alertas_insumos,
        'incidencias_recientes': incidencias_recientes,
        'dudas_recientes': dudas_recientes,

        # Servicios activos
        'servicios_totales': servicios_totales,
        'servicios_pendientes': servicios_pendientes,
        'servicios_proceso': servicios_proceso,
        'servicios_listos': servicios_listos,
        'maquinas_lavado': maquinas_lavado,
        'maquinas_secado': maquinas_secado,

        # Precios
        'prendas_destacadas': prendas_destacadas,
    }

    return render(request, 'admin/dashboard.html', context)


@solo_admin
def admin_finanzas(request):
    from django.db.models import F
    from calendar import monthrange

    hoy = timezone.now().date()

    # Determinar el periodo de filtro
    filtro = request.GET.get('filtro', 'hoy')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    tab_activa = request.GET.get('tab', 'resumen')

    if filtro == 'hoy':
        fecha_inicio = hoy
        fecha_fin = hoy
    elif filtro == 'semana':
        fecha_inicio = hoy - timedelta(days=7)
        fecha_fin = hoy
    elif filtro == 'mes':
        fecha_inicio = hoy.replace(day=1)
        fecha_fin = hoy
    elif filtro == 'personalizado' and fecha_desde and fecha_hasta:
        fecha_inicio = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    else:
        fecha_inicio = hoy
        fecha_fin = hoy

    # Calcular dias del periodo
    dias_periodo = (fecha_fin - fecha_inicio).days + 1

    # ========== DATOS FINANCIEROS - INGRESOS ==========
    pedidos_periodo = Pedido.objects.filter(
        fecha_recepcion__date__gte=fecha_inicio,
        fecha_recepcion__date__lte=fecha_fin,
        estado_pago='pagado'
    )

    ingresos_totales = pedidos_periodo.aggregate(
        total=Sum('total'))['total'] or Decimal('0')

    total_pedidos = pedidos_periodo.count()
    promedio_pedido = ingresos_totales / \
        total_pedidos if total_pedidos > 0 else Decimal('0')

    # ========== METODOS DE PAGO ==========
    pago_efectivo = pedidos_periodo.filter(metodo_pago='efectivo').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    pago_tarjeta = pedidos_periodo.filter(metodo_pago='tarjeta').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    pago_transferencia = pedidos_periodo.filter(metodo_pago='transferencia').aggregate(
        total=Sum('total'))['total'] or Decimal('0')

    total_pagos = pago_efectivo + pago_tarjeta + pago_transferencia
    pct_efectivo = round((pago_efectivo / total_pagos * 100),
                         1) if total_pagos > 0 else 0
    pct_tarjeta = round((pago_tarjeta / total_pagos * 100),
                        1) if total_pagos > 0 else 0
    pct_transferencia = round(
        (pago_transferencia / total_pagos * 100), 1) if total_pagos > 0 else 0

    # ========== GRAFICA DE PRENDAS ==========
    detalles_periodo = DetallePedido.objects.filter(
        pedido__fecha_recepcion__date__gte=fecha_inicio,
        pedido__fecha_recepcion__date__lte=fecha_fin,
        pedido__estado_pago='pagado'
    )

    prendas_stats = detalles_periodo.values(
        'prenda__nombre'
    ).annotate(
        cantidad_total=Sum('cantidad'),
        ganancia_total=Sum('subtotal')
    ).order_by('-cantidad_total')[:10]

    total_prendas = sum(p['cantidad_total']
                        for p in prendas_stats if p['cantidad_total']) if prendas_stats else 0
    prendas_data = []
    for prenda in prendas_stats:
        if prenda['prenda__nombre'] and prenda['cantidad_total']:
            pct = round(
                (prenda['cantidad_total'] / total_prendas * 100), 1) if total_prendas > 0 else 0
            prendas_data.append({
                'nombre': prenda['prenda__nombre'],
                'cantidad': prenda['cantidad_total'],
                'ganancia': float(prenda['ganancia_total'] or 0),
                'porcentaje': pct
            })

    # ========== GRAFICA DE SERVICIOS ==========
    servicios_stats = pedidos_periodo.values(
        'tipo_servicio'
    ).annotate(
        cantidad=Count('id'),
        ganancia_total=Sum('total')
    ).order_by('-cantidad')

    total_servicios = sum(s['cantidad']
                          for s in servicios_stats) if servicios_stats else 0
    servicios_data = []
    for servicio in servicios_stats:
        pct = round((servicio['cantidad'] / total_servicios *
                    100), 1) if total_servicios > 0 else 0
        servicios_data.append({
            'nombre': servicio['tipo_servicio'] or 'Sin especificar',
            'cantidad': servicio['cantidad'],
            'ganancia': float(servicio['ganancia_total'] or 0),
            'porcentaje': pct
        })

    # ========== DATOS DE GASTOS ==========
    mes_actual = hoy.month
    anio_actual = hoy.year

    # Nombres de los meses
    meses_nombres = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                     'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    meses = [{'num': i, 'nombre': meses_nombres[i-1]} for i in range(1, 13)]
    mes_nombre = meses_nombres[mes_actual - 1]

    # Obtener renta actual del mes
    renta_actual = GastoRenta.objects.filter(
        mes=mes_actual, anio=anio_actual).first()
    total_renta_mes = float(renta_actual.monto_mensual) if renta_actual else 0

    # Obtener servicios del mes
    tipos_servicio_data = [
        {'tipo': 'agua', 'nombre': 'Agua', 'icon': 'A'},
        {'tipo': 'luz', 'nombre': 'Luz', 'icon': 'L'},
        {'tipo': 'gas', 'nombre': 'Gas', 'icon': 'G'},
        {'tipo': 'internet', 'nombre': 'Internet', 'icon': 'I'},
    ]

    gastos_servicios_mes = GastoServicio.objects.filter(
        mes=mes_actual, anio=anio_actual)
    total_servicios_mes = Decimal('0')
    gastos_servicios_data = []

    for ts in tipos_servicio_data:
        gasto = gastos_servicios_mes.filter(tipo=ts['tipo']).first()
        ts['monto_actual'] = float(gasto.monto_mensual) if gasto else 0
        total_servicios_mes += Decimal(str(ts['monto_actual']))
        if ts['monto_actual'] > 0:
            gastos_servicios_data.append({
                'tipo': ts['nombre'],
                'monto': ts['monto_actual']
            })

    # Obtener empleados y sus salarios
    empleados = Usuario.objects.filter(rol__in=['admin', 'operador'])
    empleados_data = []
    total_sueldos_semanal = Decimal('0')
    sueldos_data = []

    for emp in empleados:
        salario = SalarioEmpleado.objects.filter(empleado=emp).first()
        salario_semanal = float(salario.salario_semanal) if salario else 0
        salario_diario = salario_semanal / 7 if salario_semanal > 0 else 0

        empleados_data.append({
            'id': emp.id,
            'username': emp.username,
            'first_name': emp.first_name,
            'last_name': emp.last_name,
            'rol': emp.rol,
            'salario_semanal': salario_semanal,
            'salario_diario': salario_diario,
        })
        total_sueldos_semanal += Decimal(str(salario_semanal))

        if salario_semanal > 0:
            sueldos_data.append({
                'nombre': emp.username,
                'rol': emp.rol,
                'sueldo_semanal': salario_semanal,
                'sueldo_diario': salario_diario
            })

    # Calcular sueldos mensuales (4 semanas)
    total_sueldos_mes = float(total_sueldos_semanal * 4)

    # Total gastos del mes
    total_gastos_mes = total_renta_mes + \
        float(total_servicios_mes) + total_sueldos_mes

    # ========== CALCULAR GASTOS PROPORCIONALES AL PERIODO ==========
    # Gastos diarios
    gasto_renta_diario = total_renta_mes / 30 if total_renta_mes > 0 else 0
    gasto_servicios_diario = float(
        total_servicios_mes) / 30 if total_servicios_mes > 0 else 0
    gasto_sueldos_diario = float(
        total_sueldos_semanal) / 7 if total_sueldos_semanal > 0 else 0

    # Gastos del periodo
    gastos_renta_periodo = gasto_renta_diario * dias_periodo
    gastos_servicios_periodo = gasto_servicios_diario * dias_periodo
    gastos_sueldos_periodo = gasto_sueldos_diario * dias_periodo
    total_gastos_periodo = gastos_renta_periodo + \
        gastos_servicios_periodo + gastos_sueldos_periodo

    # Utilidad neta
    utilidad_neta = float(ingresos_totales) - total_gastos_periodo

    # JSON para gastos
    gastos_json = {
        'renta': total_renta_mes,
        'servicios': gastos_servicios_data,
        'sueldos': total_sueldos_mes
    }

    context = {
        'filtro': filtro,
        'tab_activa': tab_activa,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'dias_periodo': dias_periodo,

        # Ingresos
        'ingresos_totales': float(ingresos_totales),
        'total_pedidos': total_pedidos,
        'promedio_pedido': float(promedio_pedido),
        'utilidad_neta': float(utilidad_neta),

        # Metodos de pago
        'pago_efectivo': float(pago_efectivo),
        'pago_tarjeta': float(pago_tarjeta),
        'pago_transferencia': float(pago_transferencia),
        'pct_efectivo': float(pct_efectivo),
        'pct_tarjeta': float(pct_tarjeta),
        'pct_transferencia': float(pct_transferencia),

        # JSON para graficas
        'prendas_json': json.dumps(prendas_data),
        'servicios_json': json.dumps(servicios_data),
        'metodos_pago_json': json.dumps([
            {'nombre': 'Efectivo', 'total': float(
                pago_efectivo), 'porcentaje': float(pct_efectivo)},
            {'nombre': 'Tarjeta', 'total': float(
                pago_tarjeta), 'porcentaje': float(pct_tarjeta)},
            {'nombre': 'Transferencia', 'total': float(
                pago_transferencia), 'porcentaje': float(pct_transferencia)},
        ]),
        'gastos_json': json.dumps(gastos_json),
        'sueldos_json': json.dumps(sueldos_data),

        # Datos para tab de gastos
        'meses': meses,
        'mes_actual': mes_actual,
        'mes_nombre': mes_nombre,
        'anio_actual': anio_actual,
        'renta_actual': renta_actual,
        'tipos_servicio': tipos_servicio_data,
        'empleados': empleados_data,

        # Totales mensuales
        'total_renta_mes': total_renta_mes,
        'total_servicios_mes': float(total_servicios_mes),
        'total_sueldos_mes': total_sueldos_mes,
        'total_gastos_mes': total_gastos_mes,

        # Gastos del periodo
        'gastos_renta_periodo': gastos_renta_periodo,
        'gastos_servicios_periodo': gastos_servicios_periodo,
        'gastos_sueldos_periodo': gastos_sueldos_periodo,
        'total_gastos_periodo': total_gastos_periodo,
    }
    return render(request, 'admin/finanzas/finanzas.html', context)


def admin_corte_caja(request):
    if not request.user.groups.filter(name='Administrador').exists():
        return redirect('tasks')

    # Obtener fecha de hoy
    hoy = timezone.now().date()

    # Verificar si ya existe un corte para hoy
    corte_existente = CorteCaja.objects.filter(
        fecha=hoy, responsable=request.user).first()

    # Si es POST, guardar el corte
    if request.method == 'POST':
        efectivo_contado = Decimal(request.POST.get('efectivo_contado', 0))
        tarjeta_terminal = Decimal(request.POST.get('tarjeta_terminal', 0))
        transferencia_banco = Decimal(
            request.POST.get('transferencia_banco', 0))
        justificacion = request.POST.get('justificacion', '')

        # Pedidos pagados del día de hoy
        pedidos_hoy = Pedido.objects.filter(
            fecha_recepcion__date=hoy,
            estado_pago='pagado'
        )

        # Calcular ventas por método de pago
        ventas_efectivo = pedidos_hoy.filter(metodo_pago='efectivo').aggregate(
            total=Sum('total'))['total'] or Decimal('0')
        ventas_tarjeta = pedidos_hoy.filter(metodo_pago='tarjeta').aggregate(
            total=Sum('total'))['total'] or Decimal('0')
        ventas_transferencia = pedidos_hoy.filter(metodo_pago='transferencia').aggregate(
            total=Sum('total'))['total'] or Decimal('0')

        total_ventas = ventas_efectivo + ventas_tarjeta + ventas_transferencia
        total_fisico = efectivo_contado + tarjeta_terminal + transferencia_banco
        diferencia = total_fisico - total_ventas

        # Crear o actualizar el corte
        if corte_existente:
            corte = corte_existente
            corte.efectivo_contado = efectivo_contado
            corte.tarjeta_terminal = tarjeta_terminal
            corte.transferencia_banco = transferencia_banco
            corte.total_fisico = total_fisico
            corte.diferencia = diferencia
            corte.justificacion = justificacion
            corte.fecha_hora_registro = timezone.now()
            messages.success(request, 'Corte de caja actualizado exitosamente')
        else:
            corte = CorteCaja(
                fecha=hoy,
                responsable=request.user,
                ventas_efectivo=ventas_efectivo,
                ventas_tarjeta=ventas_tarjeta,
                ventas_transferencia=ventas_transferencia,
                total_ventas=total_ventas,
                efectivo_contado=efectivo_contado,
                tarjeta_terminal=tarjeta_terminal,
                transferencia_banco=transferencia_banco,
                total_fisico=total_fisico,
                diferencia=diferencia,
                justificacion=justificacion
            )
            messages.success(request, 'Corte de caja guardado exitosamente')

        corte.save()

        # Registrar movimiento del operador
        MovimientoOperador.objects.create(
            operador=request.user,
            accion='actualizo',
            detalles=f'Corte de caja - Diferencia: ${diferencia}'
        )

        return redirect('admin_corte_caja')

    # Pedidos pagados del día de hoy
    pedidos_hoy = Pedido.objects.filter(
        fecha_recepcion__date=hoy,
        estado_pago='pagado'
    )

    # Calcular ventas por método de pago
    ventas_efectivo = pedidos_hoy.filter(metodo_pago='efectivo').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    ventas_tarjeta = pedidos_hoy.filter(metodo_pago='tarjeta').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    ventas_transferencia = pedidos_hoy.filter(metodo_pago='transferencia').aggregate(
        total=Sum('total'))['total'] or Decimal('0')

    total_ventas = ventas_efectivo + ventas_tarjeta + ventas_transferencia

    # Si existe un corte, usar esos datos
    if corte_existente:
        efectivo_contado = corte_existente.efectivo_contado
        tarjeta_terminal = corte_existente.tarjeta_terminal
        transferencia_banco = corte_existente.transferencia_banco
        total_fisico = corte_existente.total_fisico
        diferencia = corte_existente.diferencia
        justificacion = corte_existente.justificacion or ''
    else:
        # Valores por defecto (vacios)
        efectivo_contado = Decimal('0')
        tarjeta_terminal = Decimal('0')
        transferencia_banco = Decimal('0')
        total_fisico = Decimal('0')
        diferencia = Decimal('0')
        justificacion = ''

    # ========== CALCULAR GASTOS DEL DIA ==========
    mes_actual = hoy.month
    anio_actual = hoy.year

    # Obtener renta del mes
    renta_actual = GastoRenta.objects.filter(
        mes=mes_actual, anio=anio_actual).first()
    gasto_renta_diario = float(
        renta_actual.monto_mensual) / 30 if renta_actual else 0

    # Obtener servicios del mes
    gastos_servicios_mes = GastoServicio.objects.filter(
        mes=mes_actual, anio=anio_actual)

    gasto_agua = gastos_servicios_mes.filter(tipo='agua').first()
    gasto_agua_diario = float(
        gasto_agua.monto_mensual) / 30 if gasto_agua else 0

    gasto_luz = gastos_servicios_mes.filter(tipo='luz').first()
    gasto_luz_diario = float(gasto_luz.monto_mensual) / 30 if gasto_luz else 0

    gasto_gas = gastos_servicios_mes.filter(tipo='gas').first()
    gasto_gas_diario = float(gasto_gas.monto_mensual) / 30 if gasto_gas else 0

    gasto_internet = gastos_servicios_mes.filter(tipo='internet').first()
    gasto_internet_diario = float(
        gasto_internet.monto_mensual) / 30 if gasto_internet else 0

    # Obtener sueldos
    empleados = Usuario.objects.filter(rol__in=['admin', 'operador'])
    empleados_sueldos = []
    total_sueldos_diario = Decimal('0')

    for emp in empleados:
        salario = SalarioEmpleado.objects.filter(empleado=emp).first()
        if salario and salario.salario_semanal > 0:
            sueldo_diario = float(salario.salario_semanal) / 7
            empleados_sueldos.append({
                'nombre': emp.username,
                'rol': 'Admin' if emp.rol == 'admin' else 'Encargado',
                'sueldo_diario': sueldo_diario
            })
            total_sueldos_diario += Decimal(str(sueldo_diario))

    gasto_sueldos_diario = float(total_sueldos_diario)

    # Total gastos del dia
    total_gastos_dia = (gasto_renta_diario + gasto_agua_diario + gasto_luz_diario +
                        gasto_gas_diario + gasto_internet_diario + gasto_sueldos_diario)

    # Utilidad del dia
    utilidad_dia = float(total_ventas) - total_gastos_dia

    context = {
        'fecha': hoy.strftime('%d/%m/%Y'),
        'fecha_hora': timezone.now().strftime('%d/%m/%Y %H:%M'),
        'ventas_efectivo': ventas_efectivo,
        'ventas_tarjeta': ventas_tarjeta,
        'ventas_transferencia': ventas_transferencia,
        'total_ventas': total_ventas,
        'efectivo_contado': efectivo_contado,
        'tarjeta_terminal': tarjeta_terminal,
        'transferencia_banco': transferencia_banco,
        'total_fisico': total_fisico,
        'diferencia': diferencia,
        'justificacion': justificacion,
        'responsable': request.user.username,
        'corte_guardado': corte_existente is not None,

        # Gastos del dia
        'gasto_renta_diario': gasto_renta_diario,
        'gasto_agua_diario': gasto_agua_diario,
        'gasto_luz_diario': gasto_luz_diario,
        'gasto_gas_diario': gasto_gas_diario,
        'gasto_internet_diario': gasto_internet_diario,
        'gasto_sueldos_diario': gasto_sueldos_diario,
        'total_gastos_dia': total_gastos_dia,
        'utilidad_dia': utilidad_dia,
        'empleados_sueldos': empleados_sueldos,
    }

    return render(request, 'admin/finanzas/corte_caja.html', context)


@solo_admin
def admin_usuarios(request):
    trabajadores = Usuario.objects.filter(rol__in=['operador', 'admin'])
    clientes = Usuario.objects.filter(rol='cliente')
    tab = request.GET.get('tab', 'clientes')

    context = {
        'trabajadores': trabajadores,
        'clientes': clientes,
        'tab': tab,
        'total_trabajadores': trabajadores.count(),
        'total_clientes': clientes.count(),
    }
    return render(request, 'admin/usuarios/usuarios.html', context)


@solo_admin
def admin_nuevo_usuario(request):
    if request.method == 'POST':
        form = RegistroUsuarioAdminForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            rol = form.cleaned_data['rol']
            user.rol = rol
            user.save()

            if rol == 'admin':
                grupo, created = Group.objects.get_or_create(
                    name='Administrador')
                user.groups.add(grupo)
            elif rol == 'operador':
                grupo, created = Group.objects.get_or_create(name='Trabajador')
                user.groups.add(grupo)

            messages.success(request, 'Usuario registrado exitosamente.')
            return redirect('admin_usuarios')
    else:
        form = RegistroUsuarioAdminForm()

    return render(request, 'admin/usuarios/nuevo_usuario.html', {'form': form})


@solo_admin
def admin_eliminar_usuario(request, usuario_id):
    try:
        usuario = Usuario.objects.get(id=usuario_id)
        if usuario != request.user:
            usuario.delete()
            messages.success(request, 'Usuario eliminado exitosamente.')
        else:
            messages.error(request, 'No puedes eliminarte a ti mismo.')
    except Usuario.DoesNotExist:
        messages.error(request, 'Usuario no encontrado.')

    return redirect('admin_usuarios')


@solo_admin
def admin_precios(request):
    prendas = Prenda.objects.filter(activo=True).order_by('nombre')
    servicios = Servicio.objects.filter(activo=True).order_by('tipo', 'nombre')
    return render(request, 'admin/precios.html', {
        'prendas': prendas,
        'servicios': servicios
    })


@solo_admin
@require_POST
def actualizar_precio_prenda(request):
    try:
        data = json.loads(request.body)
        prenda = get_object_or_404(Prenda, id=data.get('id'))
        prenda.precio = data.get('precio')
        prenda.save()
        return JsonResponse({'success': True, 'mensaje': 'Precio actualizado correctamente'})
    except Exception as e:
        return JsonResponse({'success': False, 'mensaje': str(e)}, status=400)


@solo_admin
@require_POST
def actualizar_precio_servicio(request):
    try:
        data = json.loads(request.body)
        servicio = get_object_or_404(Servicio, id=data.get('id'))
        servicio.precio = data.get('precio')
        servicio.save()
        return JsonResponse({'success': True, 'mensaje': 'Precio actualizado correctamente'})
    except Exception as e:
        return JsonResponse({'success': False, 'mensaje': str(e)}, status=400)


@solo_admin
@require_POST
def agregar_prenda(request):
    try:
        data = json.loads(request.body)
        nombre = data.get('nombre')
        if Prenda.objects.filter(nombre=nombre).exists():
            return JsonResponse({'success': False, 'mensaje': 'Ya existe una prenda con ese nombre'}, status=400)

        prenda = Prenda.objects.create(
            nombre=nombre, precio=data.get('precio'))
        return JsonResponse({
            'success': True,
            'mensaje': 'Prenda agregada correctamente',
            'prenda': {'id': prenda.id, 'nombre': prenda.nombre, 'precio': str(prenda.precio)}
        })
    except Exception as e:
        return JsonResponse({'success': False, 'mensaje': str(e)}, status=400)


@solo_admin
@require_POST
def agregar_servicio(request):
    try:
        data = json.loads(request.body)
        nombre = data.get('nombre')
        if Servicio.objects.filter(nombre=nombre).exists():
            return JsonResponse({'success': False, 'mensaje': 'Ya existe un servicio con ese nombre'}, status=400)

        servicio = Servicio.objects.create(
            nombre=nombre,
            precio=data.get('precio'),
            tipo=data.get('tipo', 'autoservicio'),
            descripcion=data.get('descripcion', '')
        )
        return JsonResponse({
            'success': True,
            'mensaje': 'Servicio agregado correctamente',
            'servicio': {'id': servicio.id, 'nombre': servicio.nombre, 'precio': str(servicio.precio)}
        })
    except Exception as e:
        return JsonResponse({'success': False, 'mensaje': str(e)}, status=400)


@solo_admin
@require_POST
def eliminar_prenda(request):
    try:
        data = json.loads(request.body)
        prenda = get_object_or_404(Prenda, id=data.get('id'))
        prenda.activo = False
        prenda.save()
        return JsonResponse({'success': True, 'mensaje': 'Prenda eliminada correctamente'})
    except Exception as e:
        return JsonResponse({'success': False, 'mensaje': str(e)}, status=400)


@solo_admin
@require_POST
def eliminar_servicio(request):
    try:
        data = json.loads(request.body)
        servicio = get_object_or_404(Servicio, id=data.get('id'))
        servicio.activo = False
        servicio.save()
        return JsonResponse({'success': True, 'mensaje': 'Servicio eliminado correctamente'})
    except Exception as e:
        return JsonResponse({'success': False, 'mensaje': str(e)}, status=400)


@solo_admin
def obtener_precios_json(request):
    prendas = list(Prenda.objects.filter(
        activo=True).values('id', 'nombre', 'precio'))
    servicios = list(Servicio.objects.filter(activo=True).values(
        'id', 'nombre', 'tipo', 'precio', 'descripcion'))

    for prenda in prendas:
        prenda['precio'] = str(prenda['precio'])
    for servicio in servicios:
        servicio['precio'] = str(servicio['precio'])

    return JsonResponse({'prendas': prendas, 'servicios': servicios})


@login_required
def buscar_clientes(request):
    query = request.GET.get('q', '').strip()

    if len(query) < 2:
        return JsonResponse({'clientes': []})

    clientes = Usuario.objects.filter(
        rol='cliente'
    ).filter(
        Q(username__icontains=query) |
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(telefono__icontains=query)
    )[:10]

    clientes_data = [
        {
            'id': c.id,
            'username': c.username,
            'nombre_completo': f"{c.first_name} {c.last_name}".strip() or c.username,
            'telefono': c.telefono or 'Sin telefono'
        }
        for c in clientes
    ]

    return JsonResponse({'clientes': clientes_data})


@solo_admin
def admin_inventarios(request):
    from .models import MovimientoInsumo

    insumos = Insumo.objects.all().order_by('-fecha_actualizacion')
    notificaciones = NotificacionStock.objects.filter(
        atendida=False).select_related('insumo', 'usuario')

    # Obtener los ultimos movimientos automaticos de insumos (ultimos 10)
    ultimos_movimientos = MovimientoInsumo.objects.filter(
        tipo_movimiento='consumo_automatico'
    ).select_related('insumo', 'pedido').order_by('-fecha')[:10]

    if request.method == 'POST':
        form = InsumoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto agregado correctamente.')
            return redirect('admin_inventarios')
    else:
        form = InsumoForm()

    return render(request, 'admin/inventario/inventarios.html', {
        'insumos': insumos,
        'form': form,
        'notificaciones': notificaciones,
        'ultimos_movimientos': ultimos_movimientos
    })


@solo_admin
def editar_insumo(request, id):
    insumo = get_object_or_404(Insumo, id=id)
    if request.method == 'POST':
        form = InsumoForm(request.POST, instance=insumo)
        if form.is_valid():
            form.save()
            NotificacionStock.objects.filter(
                insumo=insumo, atendida=False).update(atendida=True)
            messages.success(
                request, 'Inventario actualizado y alertas resueltas.')
            return redirect('admin_inventarios')
        else:
            errores = form.errors.as_text()
            messages.error(request, f'Error al guardar: {errores}')
    return redirect('admin_inventarios')


@solo_admin
def eliminar_insumo(request, id):
    insumo = get_object_or_404(Insumo, id=id)
    insumo.delete()
    messages.success(request, 'Producto eliminado.')
    return redirect('admin_inventarios')


@solo_admin
def admin_detalles_inventario(request):
    insumos = Insumo.objects.all().order_by('nombre')
    return render(request, 'admin/inventario/detalles_inventario.html', {'insumos': insumos})


@solo_admin
def admin_historialVentas(request):
    ventas = Pedido.objects.select_related(
        'cliente', 'servicio').order_by('-fecha_recepcion')

    busqueda = request.GET.get('buscar', '').strip()
    if busqueda:
        ventas = ventas.filter(
            Q(folio__icontains=busqueda) |
            Q(cliente__username__icontains=busqueda) |
            Q(cliente__first_name__icontains=busqueda)
        )

    context = {
        'ventas': ventas,
        'total_ventas': ventas.count(),
        'busqueda': busqueda
    }
    return render(request, 'admin/historial/historial-ventas.html', context)


@solo_admin
def admin_historialMovimientos(request):
    movimientos = MovimientoOperador.objects.select_related(
        'operador', 'pedido'
    ).order_by('-fecha')

    operadores = Usuario.objects.filter(
        rol__in=['operador', 'admin']).order_by('username')

    context = {
        'movimientos': movimientos,
        'total_movimientos': movimientos.count(),
        'operadores': operadores
    }
    return render(request, 'admin/historial/historial-movimientos.html', context)


@solo_admin
def admin_detalleVenta(request, pedido_id=None):
    pedido = None
    if pedido_id:
        pedido = get_object_or_404(Pedido, id=pedido_id)

    context = {
        'pedido': pedido
    }
    return render(request, 'admin/historial/detalle-venta.html', context)


@solo_admin
def admin_incidencias(request):
    if request.method == 'POST':
        tipo = request.POST.get('tipo', 'duda')
        accion = request.POST.get('accion')

        if tipo == 'duda':
            duda_id = request.POST.get('duda_id')
            respuesta = request.POST.get('respuesta')
            try:
                duda = DudaQueja.objects.get(id=duda_id)
                if accion == 'resolver':
                    duda.respuesta = respuesta
                    duda.estado = 'resuelto'
                    duda.fecha_resolucion = timezone.now()
                    duda.save()
                    return JsonResponse({'success': True, 'message': 'Duda/queja resuelta.'})
                elif accion == 'en_proceso':
                    duda.estado = 'en_proceso'
                    duda.save()
                    return JsonResponse({'success': True, 'message': 'Actualización en proceso.'})
            except DudaQueja.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Duda/queja no encontrada.'})

        elif tipo == 'incidencia':
            incidencia_id = request.POST.get('incidencia_id')
            respuesta = request.POST.get('respuesta')
            try:
                incidencia = Incidencia.objects.get(id=incidencia_id)
                if accion == 'resolver':
                    incidencia.respuesta = respuesta
                    incidencia.estado = 'resuelto'
                    incidencia.fecha_resolucion = timezone.now()
                    incidencia.save()
                    return JsonResponse({'success': True, 'message': 'Incidencia resuelta.'})
                elif accion == 'en_proceso':
                    incidencia.estado = 'en_proceso'
                    incidencia.save()
                    return JsonResponse({'success': True, 'message': 'Incidencia en proceso.'})
            except Incidencia.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Incidencia no encontrada.'})

    dudas_quejas = DudaQueja.objects.select_related(
        'cliente').all().order_by('-fecha_creacion')
    incidencias_list = Incidencia.objects.select_related(
        'trabajador').all().order_by('-fecha_reporte')

    context = {
        'dudas_quejas': dudas_quejas,
        'incidencias': incidencias_list
    }
    return render(request, 'admin/incidencias.html', context)


@solo_admin
def admin_configuracion(request):
    incidencias_pendientes = Incidencia.objects.exclude(
        estado='resuelto').count()
    productos_bajo_stock = 0
    for insumo in Insumo.objects.all():
        if insumo.capacidad_maxima > 0:
            porcentaje = (insumo.stock_actual / insumo.capacidad_maxima) * 100
            if porcentaje <= 10:
                productos_bajo_stock += 1

    context = {
        'incidencias_pendientes': incidencias_pendientes,
        'productos_bajo_stock': productos_bajo_stock,
    }

    return render(request, 'admin/configuracion.html', context)


@solo_admin
def guardar_gasto_renta(request):
    """Guarda o actualiza el gasto de renta mensual"""
    if request.method == 'POST':
        monto = request.POST.get('monto_renta', 0)
        mes = int(request.POST.get('mes_renta', timezone.now().month))
        anio = int(request.POST.get('anio_renta', timezone.now().year))

        try:
            monto = Decimal(str(monto))

            # Crear o actualizar el registro de renta
            renta, created = GastoRenta.objects.update_or_create(
                mes=mes,
                anio=anio,
                defaults={'monto_mensual': monto}
            )

            if created:
                messages.success(
                    request, f'Renta de {mes}/{anio} registrada exitosamente: ${monto}')
            else:
                messages.success(
                    request, f'Renta de {mes}/{anio} actualizada exitosamente: ${monto}')

        except (ValueError, InvalidOperation) as e:
            messages.error(request, f'Error al guardar la renta: {str(e)}')

    return redirect('admin_finanzas')


@solo_admin
def guardar_gasto_servicio(request):
    """Guarda o actualiza el gasto de un servicio (agua, luz, gas, internet)"""
    if request.method == 'POST':
        tipo = request.POST.get('tipo_servicio')
        monto = request.POST.get('monto_servicio', 0)
        mes = int(request.POST.get('mes_servicio', timezone.now().month))
        anio = timezone.now().year

        try:
            monto = Decimal(str(monto))

            # Crear o actualizar el registro del servicio
            servicio, created = GastoServicio.objects.update_or_create(
                tipo=tipo,
                mes=mes,
                anio=anio,
                defaults={'monto_mensual': monto}
            )

            nombre_servicio = dict(
                GastoServicio.TIPOS_SERVICIO).get(tipo, tipo)

            if created:
                messages.success(
                    request, f'{nombre_servicio} de {mes}/{anio} registrado exitosamente: ${monto}')
            else:
                messages.success(
                    request, f'{nombre_servicio} de {mes}/{anio} actualizado exitosamente: ${monto}')

        except (ValueError, InvalidOperation) as e:
            messages.error(request, f'Error al guardar el servicio: {str(e)}')

    return redirect('admin_finanzas')


@solo_admin
def guardar_salario_empleado(request):
    """Guarda o actualiza el salario semanal de un empleado"""
    if request.method == 'POST':
        empleado_id = request.POST.get('empleado_id')
        salario = request.POST.get('salario_semanal', 0)

        try:
            salario = Decimal(str(salario))
            empleado = Usuario.objects.get(id=empleado_id)

            # Crear o actualizar el registro del salario
            salario_obj, created = SalarioEmpleado.objects.update_or_create(
                empleado=empleado,
                defaults={'salario_semanal': salario}
            )

            if created:
                messages.success(
                    request, f'Salario de {empleado.username} registrado exitosamente: ${salario}/semana')
            else:
                messages.success(
                    request, f'Salario de {empleado.username} actualizado exitosamente: ${salario}/semana')

        except Usuario.DoesNotExist:
            messages.error(request, 'Empleado no encontrado')
        except (ValueError, InvalidOperation) as e:
            messages.error(request, f'Error al guardar el salario: {str(e)}')

    return redirect('admin_finanzas')

# ==========================================
#              VISTAS TRABAJADOR
# ==========================================


@solo_trabajador
def trabajador_dashboard(request):
    return render(request, 'trabajador/dashboard.html')


@solo_trabajador
def servicios_proceso(request):
    pedidos = Pedido.objects.exclude(
        estado__in=['entregado', 'cancelado']
    ).select_related('cliente').order_by('fecha_recepcion')

    busqueda = request.GET.get('buscar', '').strip()
    if busqueda:
        pedidos = pedidos.filter(
            Q(folio__icontains=busqueda) |
            Q(cliente__username__icontains=busqueda) |
            Q(cliente__first_name__icontains=busqueda) |
            Q(cliente__last_name__icontains=busqueda)
        )

    return render(request, 'trabajador/procedimiento/servicios_proceso.html', {
        'pedidos': pedidos,
        'busqueda': busqueda
    })


@solo_trabajador
def historial_servicios(request):
    pedidos = Pedido.objects.filter(
        estado__in=['entregado', 'cancelado']
    ).select_related('cliente').order_by('-fecha_entrega_real', '-fecha_recepcion')

    busqueda = request.GET.get('buscar', '').strip()
    if busqueda:
        pedidos = pedidos.filter(
            Q(folio__icontains=busqueda) |
            Q(cliente__username__icontains=busqueda) |
            Q(cliente__first_name__icontains=busqueda)
        )

    return render(request, 'trabajador/procedimiento/historial_servicios.html', {
        'pedidos': pedidos
    })


@solo_trabajador
def nuevo_servicio(request):
    clientes = Usuario.objects.filter(rol='cliente').order_by('username')
    servicios = Servicio.objects.filter(activo=True)
    prendas = Prenda.objects.filter(activo=True)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            cliente_id = data.get('cliente_id')
            cliente = Usuario.objects.filter(
                id=cliente_id, rol='cliente').first()
            if not cliente:
                return JsonResponse({'success': False, 'message': 'Cliente no encontrado'}, status=400)

            pedido = Pedido.objects.create(
                cliente=cliente,
                operador=request.user,
                tipo_servicio=data.get('tipo_servicio', 'por_encargo'),
                peso=Decimal(str(data.get('peso', 0))),
                cantidad_prendas=int(data.get('cantidad_prendas', 0)),
                observaciones=data.get('observaciones', ''),
                cobija_tipo=data.get('cobija_tipo', ''),
                lavado_especial=data.get('lavado_especial', False),
                total=Decimal(str(data.get('total', 0))),
                metodo_pago=data.get('metodo_pago', 'efectivo'),
                estado='pendiente',
                estado_pago='pendiente',
                origen='operador',
                fecha_entrega_estimada=data.get(
                    'fecha_entrega') if data.get('fecha_entrega') else None
            )

            MovimientoOperador.objects.create(
                operador=request.user,
                accion='registro_servicio',
                detalles=pedido.folio,
                pedido=pedido
            )

            # Crear detalles del pedido con las prendas si vienen en los datos
            prendas_data = data.get('prendas', [])
            for prenda_data in prendas_data:
                prenda_obj = Prenda.objects.filter(
                    id=prenda_data.get('prenda_id')).first()
                if prenda_obj:
                    DetallePedido.objects.create(
                        pedido=pedido,
                        prenda=prenda_obj,
                        cantidad=prenda_data.get('cantidad', 1),
                        peso=Decimal(str(prenda_data.get('peso', 0))),
                        precio_unitario=Decimal(
                            str(prenda_data.get('precio', 0))),
                        subtotal=Decimal(str(prenda_data.get('subtotal', 0)))
                    )

            # Descontar insumos automaticamente del inventario
            resultado_descuento = descontar_insumos_por_pedido(pedido)

            pedido.refresh_from_db()
            mensaje_ticket = ""
            ticket_url = ""

            try:
                # CAMBIO CLAVE: Pasamos request para que el QR sea dinámico
                pdf_bytes = render_pdf_ticket(request, pedido)
                if pdf_bytes:
                    ticket_url = reverse('imprimir_ticket', args=[pedido.id])
                    enviado = enviar_ticket_email(pedido, pdf_bytes)
                    if enviado:
                        mensaje_ticket = " y ticket enviado por correo."
                    else:
                        mensaje_ticket = " (Ticket generado, pero no se pudo enviar correo)."
                else:
                    mensaje_ticket = " (Nota: Error al generar el PDF visual)."
            except Exception as e:
                print(f"Error en proceso de ticket: {e}")
                mensaje_ticket = " (Error técnico con el ticket PDF)."

            return JsonResponse({
                'success': True,
                'message': f'Servicio registrado correctamente{mensaje_ticket}',
                'folio': pedido.folio,
                'ticket_url': ticket_url,
                'insumos_descontados': resultado_descuento.get('descuentos', [])
            })

        except Exception as e:
            return JsonResponse({'success': False, 'message': f"Error interno: {str(e)}"}, status=400)

    # RUTA ACTUALIZADA
    return render(request, 'trabajador/servicio/nuevo_servicio.html', {
        'clientes': clientes,
        'servicios': servicios,
        'prendas': prendas
    })


@solo_trabajador
def validar_ticket(request):
    return render(request, 'trabajador/tickets/validar_ticket.html')


@solo_trabajador
def api_buscar_pedido(request):
    folio = request.GET.get('folio', '').strip().upper()
    try:
        pedido = Pedido.objects.get(folio__iexact=folio)
        data = {
            'success': True,
            'pedido': {
                'id': pedido.id,
                'folio': pedido.folio,
                'cliente': f"{pedido.cliente.first_name} {pedido.cliente.last_name}" if pedido.cliente.first_name else pedido.cliente.username,
                'servicio': pedido.tipo_servicio,
                'total': str(pedido.total),
                'estado': pedido.estado,
                'estado_display': pedido.get_estado_display(),
                'estado_pago': pedido.estado_pago,
                'peso': str(pedido.peso)
            }
        }
    except Pedido.DoesNotExist:
        data = {'success': False,
                'message': 'No existe ningún pedido con ese folio.'}
    return JsonResponse(data)


@solo_trabajador
def api_entregar_pedido(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        pedido = get_object_or_404(Pedido, id=data.get('pedido_id'))

        if pedido.estado != 'listo':
            return JsonResponse({'success': False, 'message': 'El pedido no está listo para entrega.'})

        pedido.estado = 'entregado'
        pedido.fecha_entrega_real = timezone.now()

        if pedido.estado_pago == 'pendiente':
            pedido.estado_pago = 'pagado'
            pedido.metodo_pago = 'efectivo'

        pedido.save()

        MovimientoOperador.objects.create(
            operador=request.user,
            accion='entrego',
            detalles=f"Entregó pedido {pedido.folio}",
            pedido=pedido
        )
        return JsonResponse({'success': True})

    return JsonResponse({'success': False}, status=400)


@solo_trabajador
def incidencias(request):
    if request.method == 'POST':
        asunto = request.POST.get('asunto')
        descripcion = request.POST.get('descripcion')
        prioridad = request.POST.get('prioridad', 'media')
        evidencia = request.FILES.get('evidencia')

        if asunto and asunto.strip() and descripcion and descripcion.strip():
            Incidencia.objects.create(
                trabajador=request.user,
                asunto=asunto.strip(),
                descripcion=descripcion.strip(),
                prioridad=prioridad,
                evidencia=evidencia
            )
            return JsonResponse({'success': True, 'message': 'Incidencia reportada exitosamente.'})
        return JsonResponse({'success': False, 'message': 'Por favor complete todos los campos requeridos.'})

    mis_incidencias = Incidencia.objects.filter(
        trabajador=request.user).order_by('-fecha_reporte')
    return render(request, 'trabajador/incidencias/incidencias.html', {'mis_incidencias': mis_incidencias})


@solo_trabajador
def inventario(request):
    insumos = Insumo.objects.all()
    if request.method == 'POST':
        producto_nombre = request.POST.get('producto_nombre')
        insumo_obj = Insumo.objects.filter(nombre=producto_nombre).first()
        if insumo_obj:
            NotificacionStock.objects.get_or_create(
                insumo=insumo_obj,
                atendida=False,
                defaults={'usuario': request.user}
            )
            messages.success(
                request, f'¡Aviso enviado al administrador sobre: {producto_nombre}!')
        return redirect('inventario')

    return render(request, 'trabajador/inventario/inventario.html', {'insumos': insumos})

@solo_trabajador
def detalle_servicio(request, pedido_id=None):
    pedido = get_object_or_404(Pedido, id=pedido_id) if pedido_id else None

    lavadoras_disp = Maquina.objects.filter(
        estado='disponible', tipo='lavadora')
    secadoras_disp = Maquina.objects.filter(
        estado='disponible', tipo='secadora')

    if request.method == 'POST' and pedido:
        try:
            data = json.loads(request.body)
            nuevo_estado = data.get('estado')
            estado_pago = data.get('estado_pago')
            notas = data.get('notas', '')
            maquina_id = data.get('maquina_id')
            tiempo_asignado = data.get('tiempo_asignado', 30)

            if nuevo_estado:
                pedido.estado = nuevo_estado
                if nuevo_estado == 'entregado':
                    pedido.fecha_entrega_real = timezone.now()

            if nuevo_estado == 'en_proceso' and maquina_id:
                try:
                    maquina = Maquina.objects.get(id=maquina_id)
                    if maquina.estado == 'disponible':
                        maquina.estado = 'ocupado'
                        maquina.pedido_actual = pedido
                        maquina.hora_inicio_uso = timezone.now()
                        maquina.tiempo_asignado = int(
                            tiempo_asignado) if tiempo_asignado else 30
                        maquina.save()

                        msg_sistema = f"\n[Sistema {timezone.now().strftime('%H:%M')}] Iniciado en {maquina.nombre} ({maquina.tiempo_asignado} min)."
                        if pedido.observaciones:
                            pedido.observaciones += msg_sistema
                        else:
                            pedido.observaciones = msg_sistema
                except Maquina.DoesNotExist:
                    pass

            if estado_pago:
                pedido.estado_pago = estado_pago
            
            # Actualizar metodo de pago si se proporciona
            metodo_pago = data.get('metodo_pago')
            if metodo_pago and metodo_pago in ['efectivo', 'tarjeta', 'transferencia']:
                pedido.metodo_pago = metodo_pago

            if notas:
                if pedido.observaciones:
                    pedido.observaciones += f"\n{notas}"
                else:
                    pedido.observaciones = notas

            pedido.save()

            MovimientoOperador.objects.create(
                operador=request.user,
                accion='actualizo',
                detalles=f"Actualizo pedido {pedido.folio} a estado: {nuevo_estado}",
                pedido=pedido
            )

            return JsonResponse({'success': True, 'message': 'Pedido actualizado correctamente'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    return render(request, 'trabajador/procedimiento/detalle_servicio.html', {
        'pedido': pedido,
        'lavadoras': lavadoras_disp,
        'secadoras': secadoras_disp
    })



@solo_trabajador
def estatus_maquina(request):
    if request.method == 'POST':
        accion = request.POST.get('accion')

        if accion == 'agregar':
            nombre = request.POST.get('nombre')
            tipo = request.POST.get('tipo')
            if nombre and tipo:
                Maquina.objects.create(nombre=nombre, tipo=tipo)
                messages.success(request, 'Máquina registrada correctamente.')

        elif accion == 'baja_definitiva':
            maquina_id = request.POST.get('maquina_id')
            Maquina.objects.filter(id=maquina_id).delete()
            messages.success(request, 'Máquina eliminada.')

        elif accion == 'reportar_mantenimiento':
            maquina_id = request.POST.get('maquina_id')
            maquina = get_object_or_404(Maquina, id=maquina_id)
            maquina.estado = 'mantenimiento'
            maquina.save()
            messages.warning(request, 'Máquina puesta en mantenimiento.')

        elif accion == 'toggle_uso':
            maquina_id = request.POST.get('maquina_id')
            maquina = get_object_or_404(Maquina, id=maquina_id)
            if maquina.estado == 'disponible':
                maquina.estado = 'ocupado'
            elif maquina.estado == 'ocupado':
                maquina.estado = 'disponible'
            maquina.save()

        elif accion == 'reactivar':
            maquina_id = request.POST.get('maquina_id')
            maquina = get_object_or_404(Maquina, id=maquina_id)
            maquina.estado = 'disponible'
            maquina.save()
            messages.success(request, 'Máquina reactivada y lista para usar.')

        return redirect('estatus_maquina')

    lavadoras = Maquina.objects.filter(tipo='lavadora').order_by('nombre')
    secadoras = Maquina.objects.filter(tipo='secadora').order_by('nombre')

    return render(request, 'trabajador/estatus/estatus_maquina.html', {
        'lavadoras': lavadoras,
        'secadoras': secadoras
    })


@solo_trabajador
@require_POST
def asignar_maquina(request):
    try:
        data = json.loads(request.body)
        pedido_id = data.get('pedido_id')
        maquina_id = data.get('maquina_id')
        tiempo = int(data.get('tiempo', 30))

        pedido = get_object_or_404(Pedido, id=pedido_id)
        maquina = get_object_or_404(Maquina, id=maquina_id)

        if maquina.estado != 'disponible':
            return JsonResponse({'success': False, 'message': 'La máquina no está disponible.'})

        maquina.estado = 'ocupado'
        maquina.pedido_actual = pedido
        maquina.hora_inicio_uso = timezone.now()
        maquina.tiempo_asignado = tiempo
        maquina.save()

        if maquina.tipo == 'lavadora':
            pedido.estado = 'en_proceso'

        pedido.save()

        return JsonResponse({'success': True, 'message': f'Máquina {maquina.nombre} asignada al folio {pedido.folio}'})

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@solo_trabajador
def imprimir_ticket(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    # CAMBIO CLAVE: Pasamos 'request' para generar el QR dinámico
    pdf_bytes = render_pdf_ticket(request, pedido)

    if not pdf_bytes:
        return HttpResponse("Error al generar el ticket", status=500)

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="ticket_{pedido.folio}.pdf"'
    return response


@solo_trabajador
@require_POST
def api_registrar_cliente_rapido(request):
    """
    Crea un cliente desde la pantalla de Nuevo Servicio sin recargar.
    Genera usuario y contraseña automática basada en el teléfono.
    """
    try:
        data = json.loads(request.body)

        nombre = data.get('nombre', '').strip()
        apellido = data.get('apellido', '').strip()
        telefono = data.get('telefono', '').strip()
        email = data.get('email', '').strip()

        # 1. Validaciones básicas
        if not nombre or not telefono:
            return JsonResponse({'success': False, 'message': 'Nombre y Teléfono son obligatorios.'})

        if Usuario.objects.filter(telefono=telefono).exists():
            return JsonResponse({'success': False, 'message': 'Ya existe un cliente con este teléfono.'})

        if email and Usuario.objects.filter(email=email).exists():
            return JsonResponse({'success': False, 'message': 'Ese correo ya está registrado.'})

        # 2. Generación automática de Username (Usamos el teléfono para que sea único)
        username = telefono

        # 3. Crear el usuario
        nuevo_cliente = Usuario.objects.create(
            username=username,
            first_name=nombre,
            last_name=apellido,
            email=email,
            telefono=telefono,
            rol='cliente',
            # Contraseña por defecto es su teléfono
            password=make_password(telefono)
        )

        return JsonResponse({
            'success': True,
            'message': 'Cliente registrado correctamente.',
            'cliente': {
                'id': nuevo_cliente.id,
                'nombre_completo': f"{nuevo_cliente.first_name} {nuevo_cliente.last_name} ({nuevo_cliente.telefono})"
            }
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ==========================================
#              VISTAS CLIENTE
# ==========================================

@solo_cliente
def cliente_dashboard(request):
    pedidos_activos = Pedido.objects.filter(
        cliente=request.user
    ).exclude(
        estado__in=['entregado', 'cancelado']
    ).order_by('-fecha_recepcion')

    pedidos_finalizados = Pedido.objects.filter(
        cliente=request.user,
        estado='entregado'
    ).order_by('-fecha_entrega_real')

    context = {
        'pedidos_activos': pedidos_activos,
        'pedidos_finalizados': pedidos_finalizados,
    }
    return render(request, 'cliente/dashboard.html', context)


@solo_cliente
def solicitar_servicio(request):
    servicios = Servicio.objects.filter(activo=True)
    return render(request, 'cliente/solicitar_servicio.html', {'servicios': servicios})


@solo_cliente
def perfil(request):
    usuario = request.user

    if request.method == 'POST':
        # 1. Recibir los datos del formulario (Nombres del HTML)
        nuevo_nombre = request.POST.get('first_name', '').strip()
        nuevo_apellido = request.POST.get('last_name', '').strip()
        nuevo_email = request.POST.get('email', '').strip()
        nueva_direccion = request.POST.get('direccion', '').strip()

        errores = False

        # 2. Actualizar Nombre y Apellido (Si escribieron algo)
        if nuevo_nombre:
            usuario.first_name = nuevo_nombre
        if nuevo_apellido:
            usuario.last_name = nuevo_apellido

        # 3. Actualizar Dirección (El dato clave que faltaba)
        usuario.direccion = nueva_direccion

        # 4. Validar y Actualizar Email (Con cuidado de duplicados)
        if nuevo_email and nuevo_email != usuario.email:
            # Revisar si alguien más ya usa ese correo
            if Usuario.objects.filter(email=nuevo_email).exclude(id=usuario.id).exists():
                messages.error(
                    request, '❌ Ese correo electrónico ya está registrado por otra persona.')
                errores = True
            else:
                usuario.email = nuevo_email
                # OJO: NO cambiamos usuario.username, ese sigue siendo el teléfono.

        # 5. Guardar cambios si todo está bien
        if not errores:
            usuario.save()
            messages.success(
                request, '✅ ¡Tu información ha sido actualizada correctamente!')
            return redirect('perfil')

        # Si hubo errores, el código sigue y recarga la página mostrando las alertas

    # --- SECCIÓN GET (Mostrar datos) ---
    ultimo_pedido = Pedido.objects.filter(
        cliente=usuario
    ).exclude(
        estado='cancelado'
    ).order_by('-fecha_recepcion').first()

    context = {
        'ultimo_pedido': ultimo_pedido,
        'fecha_registro': usuario.date_joined
    }
    return render(request, 'cliente/perfil.html', context)


@solo_cliente
def rastrear_servicio(request):
    return render(request, 'cliente/rastrear_servicio.html')


@solo_cliente
def dudas_quejas(request):
    if request.method == 'POST':
        comentario = request.POST.get('comentario')
        if comentario and comentario.strip():
            DudaQueja.objects.create(
                cliente=request.user, comentario=comentario.strip())
            return JsonResponse({'success': True, 'message': 'Tu comentario ha sido enviado exitosamente.'})
        return JsonResponse({'success': False, 'message': 'El comentario no puede estar vacío.'})

    mis_dudas = DudaQueja.objects.filter(
        cliente=request.user).order_by('-fecha_creacion')
    return render(request, 'cliente/dudas_quejas.html', {'mis_dudas': mis_dudas})


@solo_cliente
def autoservicio(request):
    servicios_autoservicio = Servicio.objects.filter(
        activo=True, tipo='autoservicio')

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            servicio_id = data.get('servicio_id')
            servicio_nombre = data.get('servicio_nombre')
            total = Decimal(str(data.get('total', 0)))
            metodo_pago = data.get('metodo_pago', 'efectivo')

            servicio = Servicio.objects.filter(
                id=servicio_id).first() if servicio_id else None

            pedido = Pedido.objects.create(
                cliente=request.user,
                servicio=servicio,
                tipo_servicio='Autoservicio' if not servicio_nombre else servicio_nombre,
                total=total,
                metodo_pago=metodo_pago,
                estado='pendiente',
                estado_pago='pendiente',
                origen='cliente'
            )
            return JsonResponse({
                'success': True,
                'message': 'Servicio registrado exitosamente',
                'folio': pedido.folio
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    return render(request, 'cliente/autoservicio.html', {'servicios': servicios_autoservicio})


@solo_cliente
def seleccionar_servicio(request):
    return render(request, 'cliente/seleccionar_servicio.html')


@solo_cliente
def servCosto(request):
    tipo_servicio = request.GET.get('tipo', 'por_encargo')
    tipos_nombres = {
        'autoservicio': 'Autoservicio',
        'por_encargo': 'Servicio por encargo',
        'a_domicilio': 'Servicio a domicilio',
        'tintoreria': 'Tintoreria',
    }
    tipo_servicio_nombre = tipos_nombres.get(tipo_servicio, 'Servicio')
    servicio = Servicio.objects.filter(activo=True, tipo=tipo_servicio).first()
    servicio_precio = servicio.precio if servicio else 0
    prendas = Prenda.objects.filter(activo=True)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            prendas_data = data.get('prendas', [])
            total = Decimal(str(data.get('total', 0)))
            metodo_pago = data.get('metodo_pago', 'efectivo')
            tipo = data.get('tipo_servicio', tipo_servicio)

            pedido = Pedido.objects.create(
                cliente=request.user,
                servicio=servicio,
                tipo_servicio=tipos_nombres.get(tipo, tipo_servicio_nombre),
                total=total,
                metodo_pago=metodo_pago,
                cantidad_prendas=sum([p.get('cantidad', 0)
                                     for p in prendas_data]),
                peso=sum([Decimal(str(p.get('peso', 0)))
                         for p in prendas_data]),
                estado='pendiente',
                estado_pago='pendiente',
                origen='cliente'
            )

            for prenda_data in prendas_data:
                prenda_obj = Prenda.objects.filter(
                    id=prenda_data.get('prenda_id')).first()
                if prenda_obj:
                    DetallePedido.objects.create(
                        pedido=pedido,
                        prenda=prenda_obj,
                        cantidad=prenda_data.get('cantidad', 1),
                        peso=Decimal(str(prenda_data.get('peso', 0))),
                        precio_unitario=Decimal(
                            str(prenda_data.get('precio', 0))),
                        subtotal=Decimal(str(prenda_data.get('subtotal', 0)))
                    )

            # Descontar insumos automaticamente del inventario
            resultado_descuento = descontar_insumos_por_pedido(pedido)

            return JsonResponse({
                'success': True,
                'message': 'Servicio registrado exitosamente',
                'folio': pedido.folio,
                'insumos_descontados': resultado_descuento.get('descuentos', [])
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    return render(request, 'cliente/servCosto.html', {
        'prendas': prendas,
        'tipo_servicio': tipo_servicio,
        'tipo_servicio_nombre': tipo_servicio_nombre,
        'servicio_precio': servicio_precio,
    })


@solo_cliente
def terminado(request):
    return render(request, 'cliente/terminado.html')


@login_required
def tasks(request):
    user = request.user

    if user.is_superuser or user.rol == 'admin' or user.groups.filter(name='Administrador').exists():
        return redirect('admin_dashboard')

    elif user.rol == 'operador' or user.groups.filter(name='Trabajador').exists():
        return redirect('trabajador_dashboard')

    else:
        return redirect('cliente_dashboard')


@solo_admin
def exportar_finanzas_excel(request):
    from datetime import datetime, timedelta
    from django.db.models import Sum, Count
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    hoy = timezone.now().date()

    filtro = request.GET.get('filtro', 'hoy')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    tab = request.GET.get('tab', 'resumen')

    if filtro == 'hoy':
        fecha_inicio = hoy
        fecha_fin = hoy
        periodo_nombre = f"Hoy - {hoy.strftime('%d/%m/%Y')}"
    elif filtro == 'semana':
        fecha_inicio = hoy - timedelta(days=7)
        fecha_fin = hoy
        periodo_nombre = "Ultima Semana"
    elif filtro == 'mes':
        fecha_inicio = hoy.replace(day=1)
        fecha_fin = hoy
        periodo_nombre = f"Este Mes - {hoy.strftime('%B %Y')}"
    elif filtro == 'personalizado' and fecha_desde and fecha_hasta:
        fecha_inicio = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        periodo_nombre = f"Del {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
    else:
        fecha_inicio = hoy
        fecha_fin = hoy
        periodo_nombre = f"Hoy - {hoy.strftime('%d/%m/%Y')}"

    # Calcular dias del periodo
    dias_periodo = (fecha_fin - fecha_inicio).days + 1

    pedidos_periodo = Pedido.objects.filter(
        fecha_recepcion__date__gte=fecha_inicio,
        fecha_recepcion__date__lte=fecha_fin,
        estado_pago='pagado'
    )

    ingresos_totales = pedidos_periodo.aggregate(
        total=Sum('total'))['total'] or Decimal('0')

    pago_efectivo = pedidos_periodo.filter(metodo_pago='efectivo').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    pago_tarjeta = pedidos_periodo.filter(metodo_pago='tarjeta').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    pago_transferencia = pedidos_periodo.filter(metodo_pago='transferencia').aggregate(
        total=Sum('total'))['total'] or Decimal('0')

    # ========== CALCULAR GASTOS DEL PERIODO ==========
    mes_actual = hoy.month
    anio_actual = hoy.year

    # Obtener renta del mes
    renta_actual = GastoRenta.objects.filter(
        mes=mes_actual, anio=anio_actual).first()
    total_renta_mes = float(renta_actual.monto_mensual) if renta_actual else 0

    # Obtener servicios del mes
    gastos_servicios_mes = GastoServicio.objects.filter(
        mes=mes_actual, anio=anio_actual)
    total_servicios_mes = Decimal('0')
    gastos_servicios_data = []

    tipos_servicio_lista = [
        {'tipo': 'agua', 'nombre': 'Agua'},
        {'tipo': 'luz', 'nombre': 'Luz'},
        {'tipo': 'gas', 'nombre': 'Gas'},
        {'tipo': 'internet', 'nombre': 'Internet'},
    ]

    for ts in tipos_servicio_lista:
        gasto = gastos_servicios_mes.filter(tipo=ts['tipo']).first()
        monto = float(gasto.monto_mensual) if gasto else 0
        total_servicios_mes += Decimal(str(monto))
        if monto > 0:
            gastos_servicios_data.append({
                'tipo': ts['nombre'],
                'monto': monto
            })

    # Obtener sueldos
    empleados = Usuario.objects.filter(rol__in=['admin', 'operador'])
    total_sueldos_semanal = Decimal('0')
    sueldos_data = []

    for emp in empleados:
        salario = SalarioEmpleado.objects.filter(empleado=emp).first()
        salario_semanal = float(salario.salario_semanal) if salario else 0
        salario_diario = salario_semanal / 7 if salario_semanal > 0 else 0

        if salario_semanal > 0:
            sueldos_data.append({
                'nombre': f"{emp.first_name} {emp.last_name}".strip() or emp.username,
                'rol': 'Administrador' if emp.rol == 'admin' else 'Encargado',
                'sueldo_semanal': salario_semanal,
                'sueldo_diario': salario_diario
            })
            total_sueldos_semanal += Decimal(str(salario_semanal))

    total_sueldos_mes = float(total_sueldos_semanal * 4)

    # Gastos proporcionales al periodo
    gasto_renta_diario = total_renta_mes / 30 if total_renta_mes > 0 else 0
    gasto_servicios_diario = float(total_servicios_mes) / 30 if total_servicios_mes > 0 else 0
    gasto_sueldos_diario = float(total_sueldos_semanal) / 7 if total_sueldos_semanal > 0 else 0

    gastos_renta_periodo = gasto_renta_diario * dias_periodo
    gastos_servicios_periodo = gasto_servicios_diario * dias_periodo
    gastos_sueldos_periodo = gasto_sueldos_diario * dias_periodo
    total_gastos_periodo = gastos_renta_periodo + gastos_servicios_periodo + gastos_sueldos_periodo

    # Utilidad neta
    utilidad_neta = float(ingresos_totales) - total_gastos_periodo

    detalles_periodo = DetallePedido.objects.filter(
        pedido__fecha_recepcion__date__gte=fecha_inicio,
        pedido__fecha_recepcion__date__lte=fecha_fin,
        pedido__estado_pago='pagado'
    )
    prendas_stats = detalles_periodo.values(
        'prenda__nombre'
    ).annotate(
        cantidad_total=Sum('cantidad'),
        ganancia_total=Sum('subtotal')
    ).order_by('-cantidad_total')

    servicios_stats = pedidos_periodo.values(
        'tipo_servicio'
    ).annotate(
        cantidad=Count('id'),
        ganancia_total=Sum('total')
    ).order_by('-cantidad')

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Financiero"

    header_fill = PatternFill(start_color="2d3748",
                              end_color="2d3748", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=14)
    subtotal_fill = PatternFill(
        start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")
    gastos_fill = PatternFill(
        start_color="fee2e2", end_color="fee2e2", fill_type="solid")
    positive_font = Font(color="28a745", bold=True)
    negative_font = Font(color="dc3545", bold=True)

    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = "REPORTE FINANCIERO - PUNTO LIMPIO"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:D2')
    cell = ws['A2']
    cell.value = periodo_nombre
    cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 4

    # ========== SECCION RESUMEN (siempre incluir o si tab es resumen) ==========
    if tab in ['resumen', 'ingresos']:
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "RESUMEN FINANCIERO"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

        row += 1
        ws[f'A{row}'] = "Concepto"
        ws[f'B{row}'] = "Monto"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Ingresos totales"
        ws[f'B{row}'] = float(ingresos_totales)
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = positive_font

        row += 1
        ws[f'A{row}'] = "Gastos totales"
        ws[f'B{row}'] = float(total_gastos_periodo)
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = negative_font

        row += 1
        ws[f'A{row}'] = "UTILIDAD NETA"
        ws[f'B{row}'] = float(utilidad_neta)
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True, color="28a745" if utilidad_neta >= 0 else "dc3545")
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'A{row}'].fill = subtotal_fill
        ws[f'B{row}'].fill = subtotal_fill

        row += 2

    # ========== SECCION INGRESOS ==========
    if tab in ['resumen', 'ingresos']:
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "DESGLOSE POR METODO DE PAGO"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

        row += 1
        ws[f'A{row}'] = "Metodo"
        ws[f'B{row}'] = "Monto"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Efectivo"
        ws[f'B{row}'] = float(pago_efectivo)
        ws[f'B{row}'].number_format = '$#,##0.00'

        row += 1
        ws[f'A{row}'] = "Tarjeta"
        ws[f'B{row}'] = float(pago_tarjeta)
        ws[f'B{row}'].number_format = '$#,##0.00'

        row += 1
        ws[f'A{row}'] = "Transferencia"
        ws[f'B{row}'] = float(pago_transferencia)
        ws[f'B{row}'].number_format = '$#,##0.00'

        row += 2

        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "ESTADISTICAS POR PRENDA"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

        row += 1
        ws[f'A{row}'] = "Prenda"
        ws[f'B{row}'] = "Cantidad"
        ws[f'C{row}'] = "Ganancia"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)

        for prenda in prendas_stats:
            if prenda['prenda__nombre']:
                row += 1
                ws[f'A{row}'] = prenda['prenda__nombre']
                ws[f'B{row}'] = prenda['cantidad_total']
                ws[f'C{row}'] = float(prenda['ganancia_total'] or 0)
                ws[f'C{row}'].number_format = '$#,##0.00'

        row += 2
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "ESTADISTICAS POR SERVICIO"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

        row += 1
        ws[f'A{row}'] = "Servicio"
        ws[f'B{row}'] = "Cantidad"
        ws[f'C{row}'] = "Ganancia"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)

        for servicio in servicios_stats:
            row += 1
            ws[f'A{row}'] = servicio['tipo_servicio'] or 'Sin especificar'
            ws[f'B{row}'] = servicio['cantidad']
            ws[f'C{row}'] = float(servicio['ganancia_total'] or 0)
            ws[f'C{row}'].number_format = '$#,##0.00'

        row += 2

    # ========== SECCION GASTOS ==========
    if tab in ['resumen', 'gastos']:
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "GASTOS DEL PERIODO"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

        row += 1
        ws[f'A{row}'] = "Concepto"
        ws[f'B{row}'] = "Monto"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Renta (proporcional)"
        ws[f'B{row}'] = float(gastos_renta_periodo)
        ws[f'B{row}'].number_format = '$#,##0.00'

        row += 1
        ws[f'A{row}'] = "Servicios (proporcional)"
        ws[f'B{row}'] = float(gastos_servicios_periodo)
        ws[f'B{row}'].number_format = '$#,##0.00'

        row += 1
        ws[f'A{row}'] = "Sueldos (proporcional)"
        ws[f'B{row}'] = float(gastos_sueldos_periodo)
        ws[f'B{row}'].number_format = '$#,##0.00'

        row += 1
        ws[f'A{row}'] = "TOTAL GASTOS"
        ws[f'B{row}'] = float(total_gastos_periodo)
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = negative_font
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'A{row}'].fill = gastos_fill
        ws[f'B{row}'].fill = gastos_fill

        row += 2

        # Detalle de servicios
        if gastos_servicios_data:
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = "DETALLE DE GASTOS POR SERVICIOS"
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

            row += 1
            ws[f'A{row}'] = "Servicio"
            ws[f'B{row}'] = "Monto Mensual"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)

            for gasto in gastos_servicios_data:
                row += 1
                ws[f'A{row}'] = gasto['tipo']
                ws[f'B{row}'] = gasto['monto']
                ws[f'B{row}'].number_format = '$#,##0.00'

            row += 2

        # Detalle de sueldos
        if sueldos_data:
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = "DETALLE DE SUELDOS"
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

            row += 1
            ws[f'A{row}'] = "Empleado"
            ws[f'B{row}'] = "Rol"
            ws[f'C{row}'] = "Sueldo Semanal"
            ws[f'D{row}'] = "Sueldo Diario"
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].font = Font(bold=True)

            for sueldo in sueldos_data:
                row += 1
                ws[f'A{row}'] = sueldo['nombre']
                ws[f'B{row}'] = sueldo['rol']
                ws[f'C{row}'] = sueldo['sueldo_semanal']
                ws[f'C{row}'].number_format = '$#,##0.00'
                ws[f'D{row}'] = sueldo['sueldo_diario']
                ws[f'D{row}'].number_format = '$#,##0.00'

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"reporte_financiero_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@solo_admin
def imprimir_reporte_finanzas(request):
    from datetime import datetime, timedelta
    from django.db.models import Sum, Count
    from django.template.loader import get_template
    from xhtml2pdf import pisa
    from io import BytesIO

    hoy = timezone.now().date()

    filtro = request.GET.get('filtro', 'hoy')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    tab = request.GET.get('tab', 'resumen')

    if filtro == 'hoy':
        fecha_inicio = hoy
        fecha_fin = hoy
    elif filtro == 'semana':
        fecha_inicio = hoy - timedelta(days=7)
        fecha_fin = hoy
    elif filtro == 'mes':
        fecha_inicio = hoy.replace(day=1)
        fecha_fin = hoy
    elif filtro == 'personalizado' and fecha_desde and fecha_hasta:
        fecha_inicio = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    else:
        fecha_inicio = hoy
        fecha_fin = hoy

    # Calcular dias del periodo
    dias_periodo = (fecha_fin - fecha_inicio).days + 1

    pedidos_periodo = Pedido.objects.filter(
        fecha_recepcion__date__gte=fecha_inicio,
        fecha_recepcion__date__lte=fecha_fin,
        estado_pago='pagado'
    )

    ingresos_totales = pedidos_periodo.aggregate(
        total=Sum('total'))['total'] or Decimal('0')

    pago_efectivo = pedidos_periodo.filter(metodo_pago='efectivo').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    pago_tarjeta = pedidos_periodo.filter(metodo_pago='tarjeta').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    pago_transferencia = pedidos_periodo.filter(metodo_pago='transferencia').aggregate(
        total=Sum('total'))['total'] or Decimal('0')

    total_pagos = pago_efectivo + pago_tarjeta + pago_transferencia
    pct_efectivo = round((pago_efectivo / total_pagos * 100),
                         1) if total_pagos > 0 else 0
    pct_tarjeta = round((pago_tarjeta / total_pagos * 100),
                        1) if total_pagos > 0 else 0
    pct_transferencia = round(
        (pago_transferencia / total_pagos * 100), 1) if total_pagos > 0 else 0

    # ========== CALCULAR GASTOS DEL PERIODO ==========
    mes_actual = hoy.month
    anio_actual = hoy.year

    # Obtener renta del mes
    renta_actual = GastoRenta.objects.filter(
        mes=mes_actual, anio=anio_actual).first()
    total_renta_mes = float(renta_actual.monto_mensual) if renta_actual else 0

    # Obtener servicios del mes
    gastos_servicios_mes = GastoServicio.objects.filter(
        mes=mes_actual, anio=anio_actual)
    total_servicios_mes = Decimal('0')
    gastos_servicios_data = []

    tipos_servicio_lista = [
        {'tipo': 'agua', 'nombre': 'Agua'},
        {'tipo': 'luz', 'nombre': 'Luz'},
        {'tipo': 'gas', 'nombre': 'Gas'},
        {'tipo': 'internet', 'nombre': 'Internet'},
    ]

    for ts in tipos_servicio_lista:
        gasto = gastos_servicios_mes.filter(tipo=ts['tipo']).first()
        monto = float(gasto.monto_mensual) if gasto else 0
        total_servicios_mes += Decimal(str(monto))
        if monto > 0:
            gastos_servicios_data.append({
                'tipo': ts['nombre'],
                'monto': monto
            })

    # Obtener sueldos
    empleados = Usuario.objects.filter(rol__in=['admin', 'operador'])
    total_sueldos_semanal = Decimal('0')
    sueldos_data = []

    for emp in empleados:
        salario = SalarioEmpleado.objects.filter(empleado=emp).first()
        salario_semanal = float(salario.salario_semanal) if salario else 0
        salario_diario = salario_semanal / 7 if salario_semanal > 0 else 0

        if salario_semanal > 0:
            sueldos_data.append({
                'nombre': f"{emp.first_name} {emp.last_name}".strip() or emp.username,
                'rol': 'Administrador' if emp.rol == 'admin' else 'Encargado',
                'sueldo_semanal': salario_semanal,
                'sueldo_diario': salario_diario
            })
            total_sueldos_semanal += Decimal(str(salario_semanal))

    total_sueldos_mes = float(total_sueldos_semanal * 4)

    # Gastos proporcionales al periodo
    gasto_renta_diario = total_renta_mes / 30 if total_renta_mes > 0 else 0
    gasto_servicios_diario = float(total_servicios_mes) / 30 if total_servicios_mes > 0 else 0
    gasto_sueldos_diario = float(total_sueldos_semanal) / 7 if total_sueldos_semanal > 0 else 0

    gastos_renta_periodo = gasto_renta_diario * dias_periodo
    gastos_servicios_periodo = gasto_servicios_diario * dias_periodo
    gastos_sueldos_periodo = gasto_sueldos_diario * dias_periodo
    total_gastos_periodo = gastos_renta_periodo + gastos_servicios_periodo + gastos_sueldos_periodo

    # Utilidad neta
    utilidad_neta = float(ingresos_totales) - total_gastos_periodo

    detalles_periodo = DetallePedido.objects.filter(
        pedido__fecha_recepcion__date__gte=fecha_inicio,
        pedido__fecha_recepcion__date__lte=fecha_fin,
        pedido__estado_pago='pagado'
    )
    prendas_stats = detalles_periodo.values(
        'prenda__nombre'
    ).annotate(
        cantidad_total=Sum('cantidad'),
        ganancia_total=Sum('subtotal')
    ).order_by('-cantidad_total')[:10]

    servicios_stats = pedidos_periodo.values(
        'tipo_servicio'
    ).annotate(
        cantidad=Count('id'),
        ganancia_total=Sum('total')
    ).order_by('-cantidad')

    context = {
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'dias_periodo': dias_periodo,
        'tab': tab,
        'ingresos_totales': ingresos_totales,
        'utilidad_neta': utilidad_neta,
        'pago_efectivo': pago_efectivo,
        'pago_tarjeta': pago_tarjeta,
        'pago_transferencia': pago_transferencia,
        'pct_efectivo': pct_efectivo,
        'pct_tarjeta': pct_tarjeta,
        'pct_transferencia': pct_transferencia,
        'prendas_stats': prendas_stats,
        'servicios_stats': servicios_stats,
        # Gastos
        'gastos_renta_periodo': gastos_renta_periodo,
        'gastos_servicios_periodo': gastos_servicios_periodo,
        'gastos_sueldos_periodo': gastos_sueldos_periodo,
        'total_gastos_periodo': total_gastos_periodo,
        'gastos_servicios_data': gastos_servicios_data,
        'sueldos_data': sueldos_data,
        'total_renta_mes': total_renta_mes,
        'total_servicios_mes': float(total_servicios_mes),
        'total_sueldos_mes': total_sueldos_mes,
    }

    template = get_template('admin/finanzas/reporte_finanzas_pdf.html')
    html = template.render(context)

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if pdf.err:
        return HttpResponse("Error al generar el PDF", status=500)

    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    filename = f"reporte_financiero_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@solo_admin
@require_POST
def enviar_reporte_email(request):
    """Vista para enviar el reporte financiero por correo electronico"""
    if not request.user.groups.filter(name='Administrador').exists():
        return JsonResponse({'success': False, 'message': 'No autorizado'}, status=403)

    try:
        # Obtener los datos del request
        data = json.loads(request.body)
        email_destino = data.get('email')
        filtro = data.get('filtro', 'hoy')
        fecha_desde = data.get('fecha_desde')
        fecha_hasta = data.get('fecha_hasta')
        tab = data.get('tab', 'resumen')

        if not email_destino:
            return JsonResponse({'success': False, 'message': 'Email requerido'}, status=400)

        # Validar formato de email
        from django.core.validators import validate_email
        from django.core.exceptions import ValidationError
        try:
            validate_email(email_destino)
        except ValidationError:
            return JsonResponse({'success': False, 'message': 'Email invalido'}, status=400)
        
        # Calcular fechas segun el filtro
        hoy = timezone.now().date()
        if filtro == 'hoy':
            fecha_inicio = hoy
            fecha_fin = hoy
        elif filtro == 'semana':
            fecha_inicio = hoy - timedelta(days=7)
            fecha_fin = hoy
        elif filtro == 'mes':
            fecha_inicio = hoy.replace(day=1)
            fecha_fin = hoy
        elif filtro == 'personalizado' and fecha_desde and fecha_hasta:
            fecha_inicio = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        else:
            fecha_inicio = hoy
            fecha_fin = hoy

        # Calcular dias del periodo
        dias_periodo = (fecha_fin - fecha_inicio).days + 1

        # Obtener datos financieros
        pedidos_periodo = Pedido.objects.filter(
            fecha_recepcion__date__gte=fecha_inicio,
            fecha_recepcion__date__lte=fecha_fin,
            estado_pago='pagado'
        )

        ingresos_totales = pedidos_periodo.aggregate(
            total=Sum('total'))['total'] or Decimal('0')

        # Metodos de pago
        pago_efectivo = pedidos_periodo.filter(metodo_pago='efectivo').aggregate(
            total=Sum('total'))['total'] or Decimal('0')
        pago_tarjeta = pedidos_periodo.filter(metodo_pago='tarjeta').aggregate(
            total=Sum('total'))['total'] or Decimal('0')
        pago_transferencia = pedidos_periodo.filter(metodo_pago='transferencia').aggregate(
            total=Sum('total'))['total'] or Decimal('0')

        total_pagos = pago_efectivo + pago_tarjeta + pago_transferencia
        pct_efectivo = round(
            (pago_efectivo / total_pagos * 100), 1) if total_pagos > 0 else 0
        pct_tarjeta = round((pago_tarjeta / total_pagos * 100),
                            1) if total_pagos > 0 else 0
        pct_transferencia = round(
            (pago_transferencia / total_pagos * 100), 1) if total_pagos > 0 else 0

        # ========== CALCULAR GASTOS DEL PERIODO ==========
        mes_actual = hoy.month
        anio_actual = hoy.year

        # Obtener renta del mes
        renta_actual = GastoRenta.objects.filter(
            mes=mes_actual, anio=anio_actual).first()
        total_renta_mes = float(renta_actual.monto_mensual) if renta_actual else 0

        # Obtener servicios del mes
        gastos_servicios_mes = GastoServicio.objects.filter(
            mes=mes_actual, anio=anio_actual)
        total_servicios_mes = Decimal('0')
        gastos_servicios_data = []

        tipos_servicio_lista = [
            {'tipo': 'agua', 'nombre': 'Agua'},
            {'tipo': 'luz', 'nombre': 'Luz'},
            {'tipo': 'gas', 'nombre': 'Gas'},
            {'tipo': 'internet', 'nombre': 'Internet'},
        ]

        for ts in tipos_servicio_lista:
            gasto = gastos_servicios_mes.filter(tipo=ts['tipo']).first()
            monto = float(gasto.monto_mensual) if gasto else 0
            total_servicios_mes += Decimal(str(monto))
            if monto > 0:
                gastos_servicios_data.append({
                    'tipo': ts['nombre'],
                    'monto': monto
                })

        # Obtener sueldos
        empleados = Usuario.objects.filter(rol__in=['admin', 'operador'])
        total_sueldos_semanal = Decimal('0')
        sueldos_data = []

        for emp in empleados:
            salario = SalarioEmpleado.objects.filter(empleado=emp).first()
            salario_semanal = float(salario.salario_semanal) if salario else 0
            salario_diario = salario_semanal / 7 if salario_semanal > 0 else 0

            if salario_semanal > 0:
                sueldos_data.append({
                    'nombre': f"{emp.first_name} {emp.last_name}".strip() or emp.username,
                    'rol': 'Administrador' if emp.rol == 'admin' else 'Encargado',
                    'sueldo_semanal': salario_semanal,
                    'sueldo_diario': salario_diario
                })
                total_sueldos_semanal += Decimal(str(salario_semanal))

        total_sueldos_mes = float(total_sueldos_semanal * 4)

        # Gastos proporcionales al periodo
        gasto_renta_diario = total_renta_mes / 30 if total_renta_mes > 0 else 0
        gasto_servicios_diario = float(total_servicios_mes) / 30 if total_servicios_mes > 0 else 0
        gasto_sueldos_diario = float(total_sueldos_semanal) / 7 if total_sueldos_semanal > 0 else 0

        gastos_renta_periodo = gasto_renta_diario * dias_periodo
        gastos_servicios_periodo = gasto_servicios_diario * dias_periodo
        gastos_sueldos_periodo = gasto_sueldos_diario * dias_periodo
        total_gastos_periodo = gastos_renta_periodo + gastos_servicios_periodo + gastos_sueldos_periodo

        # Utilidad neta
        utilidad_neta = float(ingresos_totales) - total_gastos_periodo

        # Datos de prendas
        detalles_periodo = DetallePedido.objects.filter(
            pedido__fecha_recepcion__date__gte=fecha_inicio,
            pedido__fecha_recepcion__date__lte=fecha_fin,
            pedido__estado_pago='pagado'
        )
        prendas_stats = detalles_periodo.values(
            'prenda__nombre'
        ).annotate(
            cantidad_total=Sum('cantidad'),
            ganancia_total=Sum('subtotal')
        ).order_by('-cantidad_total')[:10]

        # Datos de servicios
        servicios_stats = pedidos_periodo.values(
            'tipo_servicio'
        ).annotate(
            cantidad=Count('id'),
            ganancia_total=Sum('total')
        ).order_by('-cantidad')

        # Preparar contexto para el PDF
        context = {
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'dias_periodo': dias_periodo,
            'tab': tab,
            'ingresos_totales': ingresos_totales,
            'utilidad_neta': utilidad_neta,
            'pago_efectivo': pago_efectivo,
            'pago_tarjeta': pago_tarjeta,
            'pago_transferencia': pago_transferencia,
            'pct_efectivo': pct_efectivo,
            'pct_tarjeta': pct_tarjeta,
            'pct_transferencia': pct_transferencia,
            'prendas_stats': prendas_stats,
            'servicios_stats': servicios_stats,
            # Gastos
            'gastos_renta_periodo': gastos_renta_periodo,
            'gastos_servicios_periodo': gastos_servicios_periodo,
            'gastos_sueldos_periodo': gastos_sueldos_periodo,
            'total_gastos_periodo': total_gastos_periodo,
            'gastos_servicios_data': gastos_servicios_data,
            'sueldos_data': sueldos_data,
            'total_renta_mes': total_renta_mes,
            'total_servicios_mes': float(total_servicios_mes),
            'total_sueldos_mes': total_sueldos_mes,
        }

        # Generar PDF
        from django.template.loader import get_template
        from io import BytesIO
        from xhtml2pdf import pisa

        template = get_template('admin/finanzas/reporte_finanzas_pdf.html')
        html = template.render(context)

        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

        if pdf.err:
            return JsonResponse({'success': False, 'message': 'Error al generar el PDF'}, status=500)

        pdf_bytes = result.getvalue()

        # Enviar email
        from django.core.mail import EmailMessage
        from django.conf import settings

        # Determinar el nombre del periodo
        if filtro == 'hoy':
            periodo_nombre = f"del dia {fecha_inicio.strftime('%d/%m/%Y')}"
        elif filtro == 'semana':
            periodo_nombre = "de la ultima semana"
        elif filtro == 'mes':
            periodo_nombre = "del mes actual"
        else:
            periodo_nombre = f"del {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"

        subject = f'Reporte Financiero Punto Limpio - {periodo_nombre}'
        body = f'''Hola,

Adjunto encontraras el reporte financiero de Punto Limpio {periodo_nombre}.

RESUMEN DEL PERIODO:
- Ingresos totales: ${float(ingresos_totales):,.2f}
- Gastos totales: ${total_gastos_periodo:,.2f}
- Utilidad neta: ${utilidad_neta:,.2f}

DESGLOSE DE GASTOS:
- Renta (proporcional): ${gastos_renta_periodo:,.2f}
- Servicios (proporcional): ${gastos_servicios_periodo:,.2f}
- Sueldos (proporcional): ${gastos_sueldos_periodo:,.2f}

METODOS DE PAGO:
- Efectivo: ${float(pago_efectivo):,.2f} ({pct_efectivo}%)
- Tarjeta: ${float(pago_tarjeta):,.2f} ({pct_tarjeta}%)
- Transferencia: ${float(pago_transferencia):,.2f} ({pct_transferencia}%)

Este reporte fue generado automaticamente por {request.user.username} el {timezone.now().strftime('%d/%m/%Y a las %H:%M')}.

Saludos,
Sistema Punto Limpio
'''

        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=settings.EMAIL_HOST_USER,
            to=[email_destino],
        )

        # Adjuntar el PDF
        filename = f"reporte_financiero_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.pdf"
        email.attach(filename, pdf_bytes, 'application/pdf')

        # Enviar el email
        email.send()

        return JsonResponse({
            'success': True,
            'message': f'Reporte enviado exitosamente a {email_destino}'
        })

    except Exception as e:
        print(f"Error enviando reporte por email: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error al enviar el reporte: {str(e)}'
        }, status=500)


@solo_admin
def imprimir_corte_caja(request):
    """
    Genera un PDF del corte de caja del día
    """
    if not request.user.groups.filter(name='Administrador').exists():
        return HttpResponse("No autorizado", status=403)

    from django.template.loader import get_template
    from xhtml2pdf import pisa
    from io import BytesIO

    # Obtener fecha de hoy
    hoy = timezone.now().date()

    # Pedidos pagados del día de hoy
    pedidos_hoy = Pedido.objects.filter(
        fecha_recepcion__date=hoy,
        estado_pago='pagado'
    )

    # Calcular ventas por método de pago
    ventas_efectivo = pedidos_hoy.filter(metodo_pago='efectivo').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    ventas_tarjeta = pedidos_hoy.filter(metodo_pago='tarjeta').aggregate(
        total=Sum('total'))['total'] or Decimal('0')
    ventas_transferencia = pedidos_hoy.filter(metodo_pago='transferencia').aggregate(
        total=Sum('total'))['total'] or Decimal('0')

    total_ventas = ventas_efectivo + ventas_tarjeta + ventas_transferencia

    # Obtener corte guardado si existe
    corte_existente = CorteCaja.objects.filter(
        fecha=hoy, responsable=request.user).first()

    if corte_existente:
        efectivo_contado = corte_existente.efectivo_contado
        tarjeta_terminal = corte_existente.tarjeta_terminal
        transferencia_banco = corte_existente.transferencia_banco
        total_fisico = corte_existente.total_fisico
        diferencia = corte_existente.diferencia
        justificacion = corte_existente.justificacion or ''
    else:
        efectivo_contado = Decimal('0')
        tarjeta_terminal = Decimal('0')
        transferencia_banco = Decimal('0')
        total_fisico = Decimal('0')
        diferencia = Decimal('0')
        justificacion = ''

    # ========== CALCULAR GASTOS DEL DIA ==========
    mes_actual = hoy.month
    anio_actual = hoy.year

    # Obtener renta del mes
    renta_actual = GastoRenta.objects.filter(
        mes=mes_actual, anio=anio_actual).first()
    gasto_renta_diario = float(
        renta_actual.monto_mensual) / 30 if renta_actual else 0

    # Obtener servicios del mes
    gastos_servicios_mes = GastoServicio.objects.filter(
        mes=mes_actual, anio=anio_actual)

    gasto_agua = gastos_servicios_mes.filter(tipo='agua').first()
    gasto_agua_diario = float(
        gasto_agua.monto_mensual) / 30 if gasto_agua else 0

    gasto_luz = gastos_servicios_mes.filter(tipo='luz').first()
    gasto_luz_diario = float(gasto_luz.monto_mensual) / 30 if gasto_luz else 0

    gasto_gas = gastos_servicios_mes.filter(tipo='gas').first()
    gasto_gas_diario = float(gasto_gas.monto_mensual) / 30 if gasto_gas else 0

    gasto_internet = gastos_servicios_mes.filter(tipo='internet').first()
    gasto_internet_diario = float(
        gasto_internet.monto_mensual) / 30 if gasto_internet else 0

    # Obtener sueldos
    empleados_pdf = Usuario.objects.filter(rol__in=['admin', 'operador'])
    total_sueldos_diario_pdf = Decimal('0')

    for emp in empleados_pdf:
        salario = SalarioEmpleado.objects.filter(empleado=emp).first()
        if salario and salario.salario_semanal > 0:
            sueldo_diario = float(salario.salario_semanal) / 7
            total_sueldos_diario_pdf += Decimal(str(sueldo_diario))

    gasto_sueldos_diario = float(total_sueldos_diario_pdf)

    # Total gastos del dia
    total_gastos_dia = (gasto_renta_diario + gasto_agua_diario + gasto_luz_diario +
                        gasto_gas_diario + gasto_internet_diario + gasto_sueldos_diario)

    # Utilidad del dia
    utilidad_dia = float(total_ventas) - total_gastos_dia

    context = {
        'fecha': hoy.strftime('%d/%m/%Y'),
        'fecha_hora': timezone.now().strftime('%d/%m/%Y %H:%M'),
        'ventas_efectivo': ventas_efectivo,
        'ventas_tarjeta': ventas_tarjeta,
        'ventas_transferencia': ventas_transferencia,
        'total_ventas': total_ventas,
        'efectivo_contado': efectivo_contado,
        'tarjeta_terminal': tarjeta_terminal,
        'transferencia_banco': transferencia_banco,
        'total_fisico': total_fisico,
        'diferencia': diferencia,
        'justificacion': justificacion,
        'responsable': request.user.username,
        # Gastos del dia
        'gasto_renta_diario': gasto_renta_diario,
        'gasto_agua_diario': gasto_agua_diario,
        'gasto_luz_diario': gasto_luz_diario,
        'gasto_gas_diario': gasto_gas_diario,
        'gasto_internet_diario': gasto_internet_diario,
        'gasto_sueldos_diario': gasto_sueldos_diario,
        'total_gastos_dia': total_gastos_dia,
        'utilidad_dia': utilidad_dia,
    }

    # Renderizar template
    template_path = 'admin/finanzas/corte_caja_pdf.html'
    template = get_template(template_path)
    html = template.render(context)

    # Crear PDF
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if pdf.err:
        return HttpResponse('Error al generar el PDF', status=500)

    # Retornar PDF
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="corte_caja_{hoy.strftime("%Y%m%d")}.pdf"'

    return response


# ==========================================
#              VISTAS PÚBLICAS (QR)
# ==========================================
@never_cache
def rastreo_qr(request, pedido_id):
    """
    Vista pública para ver el estado del pedido al escanear el QR.
    No requiere login para facilitar el acceso rápido al cliente.
    """
    from django.shortcuts import get_object_or_404

    # Buscamos el pedido
    pedido = get_object_or_404(Pedido, id=pedido_id)

    context = {
        'pedido': pedido,
    }
    # Asegúrate de tener esta plantilla creada en gestion/templates/cliente/rastrear_servicio.html
    return render(request, 'cliente/rastrear_servicio.html', context)


@never_cache
def buscar_pedido_rastreo(request):
    """
    Vista LÓGICA para procesar la búsqueda manual por FOLIO (Texto).
    Recibe el folio del formulario, busca el ID y redirige a la vista principal.
    """
    if request.method == 'GET':
        folio_query = request.GET.get('folio', '').strip()

        if folio_query:
            # Intentar encontrar el pedido por el Folio (ignorando mayúsculas/minúsculas)
            pedido = Pedido.objects.filter(folio__iexact=folio_query).first()

            if pedido:
                # Si existe, redirigimos a la vista oficial usando su ID
                return redirect('rastreo_qr', pedido_id=pedido.id)
            else:
                # Si no existe, volvemos a la página con un mensaje de error
                messages.error(
                    request, f'No se encontró ningún servicio con el folio "{folio_query}".')

    # Si entra aquí sin buscar o si falló, renderiza la página "vacía" (buscador)
    return render(request, 'cliente/rastrear_servicio.html')

@solo_admin
def exportar_historial_ventas_excel(request):
    """Exportar historial de ventas a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # Obtener parametros de busqueda
    search = request.GET.get('search', '')

    # Obtener ventas
    ventas = Pedido.objects.select_related('cliente').order_by('-fecha_recepcion')

    if search:
        ventas = ventas.filter(
            Q(folio__icontains=search) |
            Q(cliente__first_name__icontains=search) |
            Q(cliente__last_name__icontains=search) |
            Q(cliente__username__icontains=search)
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de Ventas"

    # Estilos
    header_fill = PatternFill(start_color="2d3748", end_color="2d3748", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Titulo
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = "HISTORIAL DE VENTAS - PUNTO LIMPIO"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:G2')
    cell = ws['A2']
    cell.value = f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    cell.alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Folio', 'Cliente', 'Tipo de Servicio', 'Total', 'Estado de Pago', 'Metodo de Pago', 'Fecha']
    row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Datos
    for venta in ventas:
        row += 1
        cliente_nombre = f"{venta.cliente.first_name} {venta.cliente.last_name}".strip()
        if not cliente_nombre:
            cliente_nombre = venta.cliente.username

        estado_pago = "Pagado" if venta.estado_pago == 'pagado' else "Pendiente"
        metodo_pago = venta.metodo_pago.capitalize() if venta.metodo_pago else "No especificado"

        data = [
            venta.folio,
            cliente_nombre,
            venta.tipo_servicio or 'Sin especificar',
            float(venta.total),
            estado_pago,
            metodo_pago,
            venta.fecha_recepcion.strftime('%d/%m/%Y %H:%M')
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col == 4:  # Total
                cell.number_format = '$#,##0.00'

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18

    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"historial_ventas_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

@solo_admin
def exportar_historial_movimientos_excel(request):
    """Exportar historial de movimientos a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # Obtener parametros de filtro
    usuario = request.GET.get('usuario', '')
    accion = request.GET.get('accion', '')
    fecha = request.GET.get('fecha', '')

    # Obtener movimientos
    movimientos = MovimientoOperador.objects.select_related('operador').order_by('-fecha')

    if usuario:
        movimientos = movimientos.filter(operador__username__icontains=usuario)

    if accion:
        movimientos = movimientos.filter(accion__icontains=accion)

    if fecha:
        hoy = timezone.now().date()
        if fecha == '24h':
            fecha_inicio = hoy - timedelta(days=1)
            movimientos = movimientos.filter(fecha__date__gte=fecha_inicio)
        elif fecha == '7d':
            fecha_inicio = hoy - timedelta(days=7)
            movimientos = movimientos.filter(fecha__date__gte=fecha_inicio)
        elif fecha == '30d':
            fecha_inicio = hoy - timedelta(days=30)
            movimientos = movimientos.filter(fecha__date__gte=fecha_inicio)

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de Movimientos"

    # Estilos
    header_fill = PatternFill(start_color="2d3748", end_color="2d3748", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Titulo
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = "HISTORIAL DE MOVIMIENTOS - PUNTO LIMPIO"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:D2')
    cell = ws['A2']
    cell.value = f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    cell.alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Fecha/Hora', 'Usuario', 'Accion', 'Detalles']
    row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Diccionario de acciones
    acciones_display = {
        'registro_servicio': 'Registro servicio',
        'entrego': 'Entrego',
        'cambio_precio': 'Cambio precio',
        'creo_ticket': 'Creo ticket',
        'elimino': 'Elimino',
        'actualizo': 'Actualizo',
    }

    # Datos
    for movimiento in movimientos:
        row += 1
        accion_display = acciones_display.get(movimiento.accion, movimiento.accion)

        data = [
            movimiento.fecha.strftime('%d/%m/%Y %H:%M'),
            movimiento.operador.username if movimiento.operador else 'Sistema',
            accion_display,
            movimiento.detalles or ''
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 50

    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"historial_movimientos_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
